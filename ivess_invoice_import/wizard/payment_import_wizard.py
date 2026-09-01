import base64
import csv
import io
from datetime import datetime

from odoo import _, fields, models
from odoo.exceptions import UserError

from .invoice_import_wizard import IvessInvoiceImportWizard

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None

# La hoja "cobranzas" del formato "aguas" repite el header "letra" (y varias
# columnas más) en dos posiciones distintas (la del recibo y la de la
# factura asociada que cancela), por lo que no se puede indexar por nombre
# de columna como hace el import de facturas: se valida y se lee por
# posición fija.
EXPECTED_HEADERS = [
    "tipo de comprobante",
    "letra",
    "pto de vta",
    "num compr",
    "fecha",
    "cod cliente",
    "importe total",
    "comprobante anulado?",
    "tipo de compr asoc",
    "letra",
    "pto de venta",
    "numero compr",
    "importe",
    "medio de pago",
    "moneda",
    "tipo de cambio",
    "caja",
    "importe del movimiento",
    "cod bco",
    "sucursal del bco",
    "importe del movimiento en moneda local",
]

# Columnas que identifican la cabecera del recibo agrupado (mismo criterio
# que GROUP_KEY_FIELDS de invoice_import_wizard.py).
GROUP_KEY_FIELDS = ("tipo_comprobante", "letra", "pto_vta", "num_compr")
HEADER_ONLY_FIELDS = ("fecha", "cod_cliente", "importe_total", "comprobante_anulado")
DETAIL_FIELDS = (
    "tipo_compr_asoc",
    "letra_asoc",
    "pto_venta_asoc",
    "numero_compr_asoc",
    "importe",
    "medio_pago",
    "moneda",
    "tipo_cambio",
    "caja",
    "importe_movimiento",
    "cod_bco",
    "sucursal_bco",
    "importe_movimiento_moneda_local",
)


def group_payment_rows(rows):
    """Agrupa renglones desnormalizados de cobranza (una fila por factura
    aplicada) en recibos (cabecera + líneas de aplicación), por
    (tipo_comprobante, letra, pto_vta, num_compr). Misma lógica que
    group_invoice_rows en invoice_import_wizard.py, adaptada a los campos
    de la hoja "cobranzas".
    """
    groups_by_key = {}
    ordered_groups = []

    for row in rows:
        missing = [f for f in GROUP_KEY_FIELDS if not row.get(f)]
        if missing:
            ordered_groups.append(
                {
                    "tipo_comprobante": row.get("tipo_comprobante") or "",
                    "letra": row.get("letra") or "",
                    "pto_vta": row.get("pto_vta") or "",
                    "num_compr": row.get("num_compr") or "",
                    "lines": [],
                    "row_numbers": [row.get("_row_number")],
                    "error": _("Fila %s: faltan datos de comprobante (%s).")
                    % (row.get("_row_number"), ", ".join(missing)),
                }
            )
            continue

        key = tuple(row[f] for f in GROUP_KEY_FIELDS)
        group = groups_by_key.get(key)
        if group is None:
            group = {f: row[f] for f in GROUP_KEY_FIELDS}
            for f in HEADER_ONLY_FIELDS:
                group[f] = row.get(f)
            group["lines"] = []
            group["row_numbers"] = []
            group["error"] = None
            groups_by_key[key] = group
            ordered_groups.append(group)

        group["row_numbers"].append(row.get("_row_number"))
        group["lines"].append({f: row.get(f) for f in DETAIL_FIELDS})

    return ordered_groups


class IvessPaymentImportWizard(models.TransientModel):
    _name = "ivess.payment.import.wizard"
    _description = "Importador de cobros de clientes desde archivo Excel"

    file = fields.Binary(string="Archivo (.xlsx)", required=True)
    filename = fields.Char(string="Nombre de archivo")
    state = fields.Selection(
        [
            ("upload", "Subir archivo"),
            ("preview", "Previsualización"),
            ("done", "Resultado"),
        ],
        default="upload",
    )
    result_line_ids = fields.One2many(
        "ivess.payment.import.result.line",
        "wizard_id",
        string="Cobros",
    )
    total_count = fields.Integer(string="Total", readonly=True)
    ok_count = fields.Integer(string="OK", readonly=True)
    error_count = fields.Integer(string="Errores", readonly=True)
    skipped_count = fields.Integer(string="Anulados (no importados)", readonly=True)

    # ------------------------------------------------------------------
    # Paso 1 -> 2: leer archivo, agrupar y validar (no escribe account.payment)
    # ------------------------------------------------------------------

    def action_preview(self):
        self.ensure_one()
        if not self.file:
            raise UserError(_("Adjuntá un archivo para importar."))
        if openpyxl is None:
            raise UserError(
                _(
                    "Falta la librería 'openpyxl' en el servidor para leer"
                    " archivos .xlsx."
                )
            )
        rows = self._read_excel_rows(base64.b64decode(self.file))
        groups = group_payment_rows(rows)
        # Ver _build_preview_cache: resuelve de una sola vez, contra todos
        # los grupos, los lookups que antes se repetían con un search()
        # propio por cada recibo/línea de aplicación.
        cache = self._build_preview_cache(groups)

        self.result_line_ids.unlink()
        vals_list = [
            self._prepare_preview_line_vals(index, group, cache)
            for index, group in enumerate(groups)
        ]
        if vals_list:
            self.env["ivess.payment.import.result.line"].create(vals_list)

        self.total_count = len(self.result_line_ids)
        self.error_count = len(self.result_line_ids.filtered("has_error"))
        self.skipped_count = len(
            self.result_line_ids.filtered(
                lambda r: not r.has_error and r.comprobante_anulado
            )
        )
        self.ok_count = self.total_count - self.error_count - self.skipped_count
        self.state = "preview"
        return self._reopen()

    def action_back_to_upload(self):
        self.ensure_one()
        self.result_line_ids.unlink()
        self.state = "upload"
        return self._reopen()

    def action_export_errors(self):
        self.ensure_one()
        lines = self.result_line_ids.filtered("has_error")
        if not lines:
            raise UserError(_("No hay errores para exportar."))
        attachment = self.env["ir.attachment"].create(
            {
                "name": "errores_import_cobros_%s.csv"
                % fields.Date.context_today(self),
                "type": "binary",
                "datas": base64.b64encode(self._build_errors_csv(lines)),
                "res_model": self._name,
                "res_id": self.id,
            }
        )
        return {
            "type": "ir.actions.act_url",
            "url": "/web/content/%s?download=true" % attachment.id,
            "target": "self",
        }

    @staticmethod
    def _build_errors_csv(lines):
        # ';' como delimitador y BOM utf-8: mismo criterio que
        # IvessInvoiceImportWizard._build_errors_csv, para que lo abra bien
        # Excel en configuración regional argentina.
        buffer = io.StringIO()
        writer = csv.writer(buffer, delimiter=";")
        writer.writerow(
            [
                _("Comprobante"),
                _("Código cliente"),
                _("Cliente"),
                _("Fecha"),
                _("Importe total"),
                _("Error"),
            ]
        )
        for line in lines:
            writer.writerow(
                [
                    line.comprobante_display,
                    line.cliente_codigo,
                    line.partner_id.display_name or "",
                    line.fecha or "",
                    line.importe_total,
                    (line.error_message or "").replace("\n", " | "),
                ]
            )
        return buffer.getvalue().encode("utf-8-sig")

    def _reopen(self):
        return {
            "type": "ir.actions.act_window",
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
        }

    # ------------------------------------------------------------------
    # Lectura y parseo del Excel
    # ------------------------------------------------------------------

    @staticmethod
    def _normalize_header(value):
        return (value or "").strip().lower()

    @staticmethod
    def _to_str(value):
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    @staticmethod
    def _to_float(value):
        return float(str(value).strip())

    @staticmethod
    def _to_date(value):
        text = str(int(value)) if isinstance(value, float) else str(value).strip()
        return datetime.strptime(text, "%Y%m%d").date()

    @staticmethod
    def _get_sheet(workbook):
        for name in workbook.sheetnames:
            if name.strip().lower() == "cobranzas":
                return workbook[name]
        return workbook.worksheets[0]

    def _read_excel_rows(self, content):
        try:
            # read_only=True: lectura en modo streaming, ver el mismo
            # comentario en IvessInvoiceImportWizard._read_excel_rows.
            workbook = openpyxl.load_workbook(
                io.BytesIO(content), read_only=True, data_only=True
            )
        except Exception as exc:  # pylint: disable=broad-except # noqa: BLE001
            raise UserError(
                _("No se pudo leer el archivo como Excel: %s") % exc
            ) from exc

        sheet = self._get_sheet(workbook)
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            raise UserError(_("La hoja '%s' del archivo está vacía.") % sheet.title)

        headers = [self._normalize_header(h) for h in header_row]
        mismatches = [
            _("columna %d: se esperaba '%s', se encontró '%s'")
            % (i + 1, expected, headers[i] if i < len(headers) else "")
            for i, expected in enumerate(EXPECTED_HEADERS)
            if (headers[i] if i < len(headers) else "") != expected
        ]
        if mismatches:
            raise UserError(
                _(
                    "El formato de columnas de la hoja '%s' no coincide con"
                    " el esperado:\n%s"
                )
                % (sheet.title, "\n".join(mismatches))
            )

        rows = []
        for row_number, raw_row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            if all(cell is None for cell in raw_row):
                continue
            rows.append(self._row_to_dict(raw_row, row_number))
        return rows

    def _row_to_dict(self, raw_row, row_number):
        def cell(index):
            return raw_row[index] if index < len(raw_row) else None

        return {
            "_row_number": row_number,
            "tipo_comprobante": self._to_str(cell(0)).upper(),
            "letra": self._to_str(cell(1)).upper(),
            "pto_vta": self._to_str(cell(2)),
            "num_compr": self._to_str(cell(3)),
            "fecha": cell(4),
            "cod_cliente": self._to_str(cell(5)),
            "importe_total": cell(6),
            "comprobante_anulado": self._to_str(cell(7)).upper() == "S",
            "tipo_compr_asoc": self._to_str(cell(8)).upper(),
            "letra_asoc": self._to_str(cell(9)).upper(),
            "pto_venta_asoc": self._to_str(cell(10)),
            "numero_compr_asoc": self._to_str(cell(11)),
            "importe": cell(12),
            "medio_pago": self._to_str(cell(13)),
            "moneda": self._to_str(cell(14)),
            "tipo_cambio": self._to_str(cell(15)),
            "caja": self._to_str(cell(16)),
            "importe_movimiento": cell(17),
            "cod_bco": self._to_str(cell(18)),
            "sucursal_bco": self._to_str(cell(19)),
            "importe_movimiento_moneda_local": cell(20),
        }

    # ------------------------------------------------------------------
    # Cache de lookups para la previsualización: mismo criterio que
    # IvessInvoiceImportWizard._build_preview_cache (ver comentario ahí):
    # reemplaza los search() repetidos por recibo/línea de aplicación por
    # un puñado de búsquedas en lote sobre todo el archivo.
    # ------------------------------------------------------------------

    def _build_preview_cache(self, groups):
        invoice_refs = set()
        for group in groups:
            if group.get("error"):
                continue
            for line in group["lines"]:
                if line["tipo_compr_asoc"] == "FC":
                    invoice_refs.add(
                        IvessInvoiceImportWizard._comprobante_ref(
                            line["letra_asoc"],
                            line["pto_venta_asoc"],
                            line["numero_compr_asoc"],
                        )
                    )

        invoice_by_ref = {}
        if invoice_refs:
            for move in self.env["account.move"].search(
                [("move_type", "=", "out_invoice"), ("ref", "in", list(invoice_refs))]
            ):
                invoice_by_ref.setdefault(move.ref, move)

        payment_mode_by_key = {}
        for mapping in self.env["ivess.payment.import.payment.method.code"].search(
            [("company_id", "=", self.env.company.id)]
        ):
            key = (
                mapping.medio_pago,
                mapping.caja,
                mapping.cod_bco,
                mapping.sucursal_bco,
            )
            payment_mode_by_key.setdefault(key, mapping.journal_id)

        receipt_journal, receipt_journal_candidates = self._find_receipt_journal()

        # detail_results se calcula una sola vez por grupo (resuelve todas
        # las líneas de aplicación contra los caches de arriba) y se
        # reutiliza tanto para el prefetch de deduplicación de abajo como
        # para el vals final en _prepare_preview_line_vals: evita repetir
        # la misma resolución dos veces.
        detail_results = {}
        for index, group in enumerate(groups):
            if group.get("error"):
                continue
            detail_results[index] = self._resolve_detail_lines(
                group["lines"], invoice_by_ref, payment_mode_by_key
            )

        dedup_keys = set()
        for index, group in enumerate(groups):
            if group.get("error") or group["comprobante_anulado"]:
                continue
            _detail_vals, _detail_errors, matched_partners, _applied_total = (
                detail_results[index]
            )
            if len(matched_partners) != 1:
                continue
            partner = next(iter(matched_partners))
            ref = IvessInvoiceImportWizard._comprobante_ref(
                group["letra"], group["pto_vta"], group["num_compr"]
            )
            dedup_keys.add((ref, partner.id))

        orders_by_key = {}
        if dedup_keys:
            refs = {k[0] for k in dedup_keys}
            partner_ids = {k[1] for k in dedup_keys}
            for order in self.env["account.payment.order"].search(
                [
                    ("type", "=", "receipt"),
                    ("reference", "in", list(refs)),
                    ("partner_id", "in", list(partner_ids)),
                ]
            ):
                okey = (order.reference, order.partner_id.id)
                orders_by_key[okey] = (
                    orders_by_key.get(okey, self.env["account.payment.order"]) | order
                )

        return {
            "receipt_journal": receipt_journal,
            "receipt_journal_candidates": receipt_journal_candidates,
            "detail_results": detail_results,
            "orders_by_key": orders_by_key,
        }

    # ------------------------------------------------------------------
    # Resolución de cada grupo contra Odoo (cliente vía la factura que
    # cancela cada línea de aplicación) y armado de los vals de la línea de
    # previsualización (se crean todas juntas en lote, ver action_preview).
    # ------------------------------------------------------------------

    def _prepare_preview_line_vals(self, index, group, cache):
        if group.get("error"):
            return {
                "wizard_id": self.id,
                "tipo_comprobante": group["tipo_comprobante"],
                "letra": group["letra"],
                "pto_vta": group["pto_vta"],
                "numero_comprobante": group["num_compr"],
                "has_error": True,
                "error_message": group["error"],
            }

        errors = []
        if group["tipo_comprobante"] != "RMP":
            errors.append(
                _(
                    "tipo de comprobante '%s' no soportado (esta versión solo"
                    " importa 'RMP')."
                )
                % group["tipo_comprobante"]
            )

        try:
            fecha = self._to_date(group["fecha"]) if group["fecha"] else False
        except ValueError:
            fecha = False
            errors.append(_("Fecha inválida: '%s'.") % group["fecha"])

        try:
            importe_total = self._to_float(group["importe_total"])
        except (TypeError, ValueError):
            importe_total = 0.0
            errors.append(_("Importe total inválido: '%s'.") % group["importe_total"])

        comprobante_ref = IvessInvoiceImportWizard._comprobante_ref(
            group["letra"], group["pto_vta"], group["num_compr"]
        )

        journal = cache["receipt_journal"]
        journal_candidates = cache["receipt_journal_candidates"]
        if not journal:
            if journal_candidates:
                errors.append(
                    _(
                        "Hay %s diarios de tipo 'Recibo de Cobranza' en la"
                        " compañía; debe haber uno solo para poder"
                        " determinarlo automáticamente."
                    )
                    % journal_candidates
                )
            else:
                errors.append(
                    _(
                        "No se encontró ningún diario de tipo 'Recibo de"
                        " Cobranza' (account.journal type='receipt') en la"
                        " compañía."
                    )
                )

        detail_vals, detail_errors, matched_partners, applied_total = cache[
            "detail_results"
        ][index]
        errors.extend(detail_errors)

        partner = False
        if not matched_partners:
            errors.append(
                _(
                    "No se pudo determinar el cliente: ninguna factura"
                    " aplicada pudo resolverse."
                )
            )
        elif len(matched_partners) > 1:
            errors.append(
                _(
                    "El recibo aplica a facturas de más de un cliente"
                    " distinto (%s); no se puede determinar un único cliente."
                )
                % ", ".join(sorted(p.display_name for p in matched_partners))
            )
        else:
            partner = list(matched_partners)[0]

        if not errors and abs(applied_total - importe_total) > 0.01:
            errors.append(
                _(
                    "La suma de los importes aplicados a facturas (%.2f) no"
                    " coincide con el importe total del recibo (%.2f)."
                )
                % (applied_total, importe_total)
            )

        if partner and not group["comprobante_anulado"]:
            candidates = cache["orders_by_key"].get(
                (comprobante_ref, partner.id), self.env["account.payment.order"]
            )
            existing = candidates[:1]
            if existing:
                errors.append(
                    _(
                        "Ya existe un cobro importado con esta clave"
                        " (account.payment.order #%s)."
                    )
                    % existing.id
                )

        return {
            "wizard_id": self.id,
            "tipo_comprobante": group["tipo_comprobante"],
            "letra": group["letra"],
            "pto_vta": group["pto_vta"],
            "numero_comprobante": group["num_compr"],
            "comprobante_ref": comprobante_ref,
            "fecha": fecha,
            "importe_total": importe_total,
            "comprobante_anulado": bool(group["comprobante_anulado"]),
            "cliente_codigo": group["cod_cliente"],
            "partner_id": partner.id if partner else False,
            "journal_id": journal.id if journal else False,
            "has_error": bool(errors),
            "error_message": "\n".join(errors) if errors else False,
            "detail_line_ids": detail_vals,
        }

    def _find_receipt_journal(self):
        """Resuelve el diario de Recibo de Cobranza (account.journal con
        type='receipt') de la compañía activa.

        A diferencia del diario de ventas de facturas (que sí tiene un
        diario Odoo distinto por letra + punto de venta AFIP, ver
        IvessInvoiceImportWizard._find_journal), un recibo de cobranza no
        es un comprobante autorizado por AFIP: no hay múltiples diarios
        'receipt' por letra/punto de venta para desambiguar (de hecho
        account.journal.denomination ni siquiera tiene un valor que
        represente la letra "X" que usa "aguas" para sus recibos, y el
        código del diario no tiene por qué tener dígitos de punto de
        venta). En la práctica hay un único diario 'receipt' por compañía,
        así que se usa ese directamente. Se resuelve una sola vez por
        previsualización (no depende de datos de la fila), ver
        _build_preview_cache.

        :return: tupla (diario o None, cantidad de diarios candidatos).
        """
        journals = self.env["account.journal"].search(
            [
                ("type", "=", "receipt"),
                ("company_id", "=", self.env.company.id),
            ]
        )
        return (journals[0] if len(journals) == 1 else None), len(journals)

    def _resolve_detail_lines(self, lines, invoice_by_ref, payment_mode_by_key):
        detail_vals = []
        errors = []
        matched_partners = set()
        applied_total = 0.0

        if not lines:
            errors.append(_("El recibo no tiene líneas de aplicación a facturas."))
            return detail_vals, errors, matched_partners, applied_total

        for line in lines:
            line_errors = []

            try:
                importe = self._to_float(line["importe"])
            except (TypeError, ValueError):
                importe = 0.0
                line_errors.append(_("Importe inválido: '%s'.") % line["importe"])

            invoice = self.env["account.move"]
            if line["tipo_compr_asoc"] != "FC":
                line_errors.append(
                    _("tipo de comprobante asociado '%s' no soportado (solo 'FC').")
                    % line["tipo_compr_asoc"]
                )
            else:
                invoice_ref = IvessInvoiceImportWizard._comprobante_ref(
                    line["letra_asoc"],
                    line["pto_venta_asoc"],
                    line["numero_compr_asoc"],
                )
                invoice = invoice_by_ref.get(invoice_ref, self.env["account.move"])
                if not invoice:
                    line_errors.append(
                        _("No se encontró la factura '%s' para conciliar.")
                        % invoice_ref
                    )
                elif invoice.state != "posted":
                    line_errors.append(
                        _("La factura '%s' no está confirmada (estado: %s).")
                        % (invoice_ref, invoice.state)
                    )
                elif importe - invoice.amount_residual > 0.01:
                    # account.payment.order.line._check_amount_over_original()
                    # (l10n_ar_eynes) rechaza con un error genérico en inglés
                    # si el importe supera el saldo pendiente: se valida acá
                    # antes para dar un error claro en la previsualización en
                    # vez de que reviente recién al confirmar.
                    line_errors.append(
                        _(
                            "El importe aplicado (%.2f) supera el saldo"
                            " pendiente de la factura '%s' (%.2f)."
                        )
                        % (importe, invoice_ref, invoice.amount_residual)
                    )
                else:
                    matched_partners.add(invoice.partner_id)
                    applied_total += importe

            payment_mode = None
            if line["medio_pago"]:
                payment_mode = payment_mode_by_key.get(
                    (
                        line["medio_pago"],
                        line["caja"],
                        line["cod_bco"],
                        line["sucursal_bco"],
                    )
                )
            if not payment_mode:
                line_errors.append(
                    _(
                        "No hay un diario mapeado para el medio de pago '%s'"
                        " + caja '%s' + cod bco '%s' + sucursal '%s'"
                        " (configurar en Contabilidad > Configuración >"
                        " Medios de pago (importación de cobros))."
                    )
                    % (
                        line["medio_pago"],
                        line["caja"],
                        line["cod_bco"],
                        line["sucursal_bco"],
                    )
                )

            try:
                importe_movimiento = self._to_float(line["importe_movimiento"])
            except (TypeError, ValueError):
                importe_movimiento = 0.0

            try:
                importe_movimiento_moneda_local = self._to_float(
                    line["importe_movimiento_moneda_local"]
                )
            except (TypeError, ValueError):
                importe_movimiento_moneda_local = 0.0

            detail_vals.append(
                (
                    0,
                    0,
                    {
                        "tipo_compr_asoc": line["tipo_compr_asoc"],
                        "letra_asoc": line["letra_asoc"],
                        "pto_venta_asoc": line["pto_venta_asoc"],
                        "numero_compr_asoc": line["numero_compr_asoc"],
                        "invoice_id": invoice.id if invoice else False,
                        "importe": importe,
                        "payment_mode_id": payment_mode.id if payment_mode else False,
                        "medio_pago": line["medio_pago"],
                        "moneda": line["moneda"],
                        "tipo_cambio": line["tipo_cambio"],
                        "caja": line["caja"],
                        "importe_movimiento": importe_movimiento,
                        "cod_bco": line["cod_bco"],
                        "sucursal_bco": line["sucursal_bco"],
                        "importe_movimiento_moneda_local": (
                            importe_movimiento_moneda_local
                        ),
                        "has_error": bool(line_errors),
                        "error_message": "; ".join(line_errors)
                        if line_errors
                        else False,
                    },
                )
            )
            errors.extend(line_errors)

        return detail_vals, errors, matched_partners, applied_total

    # ------------------------------------------------------------------
    # Paso 2 -> 3: crear los cobros que no tengan error
    # ------------------------------------------------------------------

    # Cada cuántos recibos se hace un commit() intermedio (ver
    # action_confirm): mismo criterio que
    # IvessInvoiceImportWizard._CONFIRM_CHUNK_SIZE, para que un import largo
    # no dependa de una única transacción gigante.
    _CONFIRM_COMMIT_EVERY = 200

    def action_confirm(self):
        self.ensure_one()
        # Ver IvessInvoiceImportWizard.action_confirm: tracking_disable evita
        # el overhead de chatter de mail.thread en cada create()/write(), que
        # no aporta nada en una importación histórica.
        self = self.with_context(tracking_disable=True)
        processed = 0
        for result_line in self.result_line_ids:
            if result_line.has_error:
                result_line.resultado = "error"
                continue
            if result_line.comprobante_anulado:
                # Un recibo anulado en el sistema origen nunca llegó a
                # cobrarse: a diferencia de una factura anulada (que sí tuvo
                # CAE y debe registrarse y cancelarse para no perder
                # numeración), acá no hay nada que registrar en Odoo.
                result_line.resultado = "skipped"
                continue
            try:
                with self.env.cr.savepoint():
                    payment_order = self._create_payment_order(result_line)
                result_line.write(
                    {"resultado": "ok", "odoo_payment_order_id": payment_order.id}
                )
            except Exception as exc:  # pylint: disable=broad-except # noqa: BLE001
                result_line.write(
                    {
                        "resultado": "error",
                        "has_error": True,
                        "error_message": _("Error al crear el cobro: %s") % exc,
                    }
                )
            processed += 1
            if processed % self._CONFIRM_COMMIT_EVERY == 0:
                # Confirma en tandas en vez de dejar TODO el import en una
                # sola transacción: si el proceso se corta a mitad de camino
                # (timeout, reinicio), lo ya procesado queda guardado (ver
                # el mismo comentario, con más detalle, en
                # IvessInvoiceImportWizard.action_confirm).
                # pylint: disable=invalid-commit
                self.env.cr.commit()

        self.ok_count = len(
            self.result_line_ids.filtered(lambda r: r.resultado == "ok")
        )
        self.error_count = len(
            self.result_line_ids.filtered(lambda r: r.resultado == "error")
        )
        self.skipped_count = len(
            self.result_line_ids.filtered(lambda r: r.resultado == "skipped")
        )
        self.state = "done"
        return self._reopen()

    def _create_payment_order(self, result_line):
        company = result_line.journal_id.company_id
        order = self.env["account.payment.order"].create(
            {
                "partner_id": result_line.partner_id.id,
                "journal_id": result_line.journal_id.id,
                "type": "receipt",
                "company_id": company.id,
                "date": result_line.fecha,
                "reference": result_line.comprobante_ref,
                "name": result_line.comprobante_ref,
                "disable_retentions": True,
            }
        )
        order.income_line_ids = [
            (0, 0, vals) for vals in self._prepare_income_line_vals(order, result_line)
        ]
        order.payment_mode_line_ids = [
            (0, 0, vals) for vals in self._prepare_payment_mode_line_vals(result_line)
        ]
        order.proforma_voucher()
        return order

    @staticmethod
    def _prepare_payment_mode_line_vals(result_line):
        """Agrupa las líneas de detalle por diario de medio de pago resuelto
        (payment_mode_id, ver _resolve_detail_lines) y arma una
        payment_mode_line por cada diario distinto, con el importe sumado
        de las facturas aplicadas con ese medio de pago: un mismo recibo
        puede combinar más de un medio de pago (ej. parte efectivo, parte
        cheque)."""
        amounts_by_journal = {}
        order_by_journal = []
        for detail in result_line.detail_line_ids:
            journal = detail.payment_mode_id
            if journal.id not in amounts_by_journal:
                amounts_by_journal[journal.id] = 0.0
                order_by_journal.append(journal)
            amounts_by_journal[journal.id] += detail.importe

        return [
            {
                "payment_mode_id": journal.id,
                "amount": amounts_by_journal[journal.id],
                "name": result_line.comprobante_ref,
                "date": result_line.fecha,
            }
            for journal in order_by_journal
        ]

    @staticmethod
    def _prepare_income_line_vals(order, result_line):
        income_vals = []
        for detail in result_line.detail_line_ids:
            invoice = detail.invoice_id
            move_line = invoice.line_ids.filtered(
                lambda aml: aml.account_id.account_type == "asset_receivable"
                and not aml.reconciled
            )[:1]
            if not move_line:
                raise UserError(
                    _(
                        "La factura '%s' no tiene un apunte a cobrar"
                        " pendiente de conciliar."
                    )
                    % invoice.display_name
                )
            income_vals.append(
                {
                    "invoice_id": invoice.id,
                    "move_line_id": move_line.id,
                    "account_id": move_line.account_id.id,
                    "amount": detail.importe,
                    "type": "income",
                    "currency_id": order.currency_id.id,
                    "original_currency_id": order.company_id.currency_id.id,
                }
            )
        return income_vals
