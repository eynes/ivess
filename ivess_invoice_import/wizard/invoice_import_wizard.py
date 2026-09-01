import base64
import csv
import io
from datetime import datetime

from odoo import _, fields, models
from odoo.exceptions import UserError

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None

EXPECTED_HEADERS = [
    "tipo de comprobante",
    "letra",
    "pto vta",
    "numero comprob",
    "fecha",
    "cod cliente",
    "razon social",
    "documento",
    "fecha vto",
    "importe total",
    "comprobante anulado",
    "cae",
    "vto cae",
    "tipo de item",
    "cod art",
    "cantidad",
    "precio unitario",
    "tasa de iva",
    "tasa de iva no inscripto",
    "importe iva inscripto",
    "importe iva no inscripto",
    "importe total neto del item",
    "importe del renglon",
    "cod impuesto interno",
    "monto imp interno",
    "cod imp especiales",
    "monto imp especiales",
    "base imp especiales",
    "cod imp especiales1",
    "monto imp especiales1",
    "base imp especiales1",
    "cod imp especiales2",
    "monto imp especiales2",
    "base imp especiales2",
]

# Columnas de impuesto especial/interno del Excel: cada tupla es (columna de
# código, columna de monto, columna de base imponible). "cod impuesto
# interno" es la única sin columna de base propia: el impuesto interno no
# tiene base imponible, así que su base se informa siempre en 0. El resto
# ("cod imp especiales", "...especiales1", "...especiales2") sí trae su
# propia base imponible en el Excel. Todas matchean igual: por código
# Bejerman (ver AccountTax.bejerman_code) o, si no matchea, por el mapeo
# manual de ivess.invoice.import.tax.code.
SPECIAL_TAX_COLUMNS = [
    ("cod impuesto interno", "monto imp interno", None),
    ("cod imp especiales", "monto imp especiales", "base imp especiales"),
    ("cod imp especiales1", "monto imp especiales1", "base imp especiales1"),
    ("cod imp especiales2", "monto imp especiales2", "base imp especiales2"),
]

# Mapeo tipo de comprobante (columna del Excel) -> res.voucher.type.doc_type,
# para resolver el comprobante Odoo (FC=Factura, ND=Nota de Débito, NC=Nota
# de Crédito) y derivar el move_type/is_debit_note correspondiente.
TIPO_COMPROBANTE_DOC_TYPES = {
    "FC": "b",
    "ND": "dn",
    "NC": "cn",
}

# Columnas que identifican la cabecera de la factura agrupada. Se toman del
# PRIMER renglón visto de cada grupo (decisión confirmada: si difieren entre
# renglones del mismo comprobante -como puede pasar con fecha vto/importe
# total en archivos reales- se ignora la diferencia y se usa el primero).
GROUP_KEY_FIELDS = ("tipo_comprobante", "letra", "pto_vta", "numero_comprob")
HEADER_ONLY_FIELDS = (
    "fecha",
    "cod_cliente",
    "razon_social",
    "documento",
    "fecha_vto",
    "importe_total",
    "comprobante_anulado",
    "cae",
    "vto_cae",
)
DETAIL_FIELDS = (
    "tipo_item",
    "cod_art",
    "cantidad",
    "precio_unitario",
    "tasa_iva",
    "tasa_iva_no_inscripto",
    "importe_iva_inscripto",
    "importe_iva_no_inscripto",
    "importe_total_neto_item",
    "importe_del_renglon",
    "special_taxes",
)


def group_invoice_rows(rows):
    """Agrupa renglones desnormalizados de factura (una fila por ítem) en
    facturas (cabecera + líneas de detalle), por (tipo_comprobante, letra,
    pto_vta, numero_comprob).

    :param rows: lista de dicts, cada uno con al menos las claves de
        GROUP_KEY_FIELDS + HEADER_ONLY_FIELDS + DETAIL_FIELDS (además puede
        traer "_row_number" para trazabilidad de errores).
    :return: lista de dicts, cada uno con las claves de GROUP_KEY_FIELDS +
        HEADER_ONLY_FIELDS + "lines" (lista de dicts con DETAIL_FIELDS) +
        "row_numbers" (números de fila origen) + "error" (str u None, para
        renglones que no se pudieron agrupar por faltarles la clave).
    """
    groups_by_key = {}
    ordered_groups = []

    for row in rows:
        missing = [f for f in GROUP_KEY_FIELDS if not row.get(f)]
        if missing:
            ordered_groups.append(
                {
                    "tipo_comprobante": row.get("tipo_comprobante") or "",
                    "letra": row.get("letra") or "",
                    "pto_vta": row.get("pto_vta") or "",
                    "numero_comprob": row.get("numero_comprob") or "",
                    "lines": [],
                    "row_numbers": [row.get("_row_number")],
                    "error": _("Fila %s: faltan datos de comprobante (%s).")
                    % (row.get("_row_number"), ", ".join(missing)),
                }
            )
            continue

        key = tuple(row[f] for f in GROUP_KEY_FIELDS)
        group = groups_by_key.get(key)
        if group is None:
            group = {f: row[f] for f in GROUP_KEY_FIELDS}
            for f in HEADER_ONLY_FIELDS:
                group[f] = row.get(f)
            group["lines"] = []
            group["row_numbers"] = []
            group["error"] = None
            groups_by_key[key] = group
            ordered_groups.append(group)

        group["row_numbers"].append(row.get("_row_number"))
        group["lines"].append({f: row.get(f) for f in DETAIL_FIELDS})

    return ordered_groups


class IvessInvoiceImportWizard(models.TransientModel):
    _name = "ivess.invoice.import.wizard"
    _description = "Importador de facturas de clientes desde archivo Excel"

    file = fields.Binary(string="Archivo (.xlsx)", required=True)
    filename = fields.Char(string="Nombre de archivo")
    state = fields.Selection(
        [
            ("upload", "Subir archivo"),
            ("preview", "Previsualización"),
            ("done", "Resultado"),
        ],
        default="upload",
    )
    result_line_ids = fields.One2many(
        "ivess.invoice.import.result.line",
        "wizard_id",
        string="Facturas",
    )
    # Dos One2many al mismo modelo/inverse pero con domain propio: a
    # diferencia del atributo "domain" de la vista (que en un widget
    # one2many embebido solo se usa para el diálogo de "agregar
    # existente", NO filtra las filas ya cargadas), un domain a nivel de
    # campo sí restringe qué filas trae el ORM al leerlo. Se usan para que
    # los botones "Ver OK"/"Ver errores" muestren cada uno una lista
    # realmente filtrada (con su propia paginación), en vez de reutilizar
    # result_line_ids (que siempre trae todo).
    result_line_ok_ids = fields.One2many(
        "ivess.invoice.import.result.line",
        "wizard_id",
        string="Facturas OK",
        domain=[("has_error", "=", False)],
    )
    result_line_error_ids = fields.One2many(
        "ivess.invoice.import.result.line",
        "wizard_id",
        string="Facturas con error",
        domain=[("has_error", "=", True)],
    )
    total_count = fields.Integer(string="Total", readonly=True)
    ok_count = fields.Integer(string="OK", readonly=True)
    error_count = fields.Integer(string="Errores", readonly=True)
    preview_show_errors = fields.Boolean(
        string="Mostrar errores",
        default=True,
        help="Controla si la previsualización muestra la lista de"
        " comprobantes OK o la de comprobantes con error.",
    )

    # ------------------------------------------------------------------
    # Paso 1 -> 2: leer archivo, agrupar y validar (no escribe account.move)
    # ------------------------------------------------------------------

    def action_preview(self):
        self.ensure_one()
        if not self.file:
            raise UserError(_("Adjuntá un archivo para importar."))
        if openpyxl is None:
            raise UserError(
                _(
                    "Falta la librería 'openpyxl' en el servidor para leer archivos .xlsx."
                )
            )

        rows = self._read_excel_rows(base64.b64decode(self.file))
        groups = group_invoice_rows(rows)
        # Se resuelven de una sola vez, contra toda la lista de grupos, los
        # lookups que en versiones anteriores se repetían con un search()
        # propio por cada factura/línea/impuesto especial (ver
        # _build_preview_cache): con archivos de miles de filas eso son
        # decenas de miles de queries individuales. El cache resuelve la
        # MISMA respuesta (mismos matches, misma ambigüedad, mismos
        # mensajes de error) con una cantidad de queries que depende de la
        # cantidad de valores DISTINTOS en el archivo, no de la cantidad de
        # filas.
        cache = self._build_preview_cache(groups)

        self.result_line_ids.unlink()
        vals_list = [self._prepare_preview_line_vals(group, cache) for group in groups]
        if vals_list:
            # Un solo create() en lote para todas las facturas del archivo,
            # en vez de un create() por factura: menos overhead de ORM
            # (flush/compute) para archivos de miles de filas.
            self.env["ivess.invoice.import.result.line"].create(vals_list)

        self.total_count = len(self.result_line_ids)
        self.error_count = len(self.result_line_ids.filtered("has_error"))
        self.ok_count = self.total_count - self.error_count
        self.preview_show_errors = True
        self.state = "preview"
        return self._reopen()

    def action_show_errors(self):
        self.ensure_one()
        self.preview_show_errors = True
        return self._reopen()

    def action_show_ok(self):
        self.ensure_one()
        self.preview_show_errors = False
        return self._reopen()

    def action_back_to_upload(self):
        self.ensure_one()
        self.result_line_ids.unlink()
        self.state = "upload"
        return self._reopen()

    def action_export_errors(self):
        self.ensure_one()
        lines = self.result_line_ids.filtered("has_error")
        if not lines:
            raise UserError(_("No hay errores para exportar."))
        attachment = self.env["ir.attachment"].create(
            {
                "name": "errores_import_facturas_%s.csv"
                % fields.Date.context_today(self),
                "type": "binary",
                "datas": base64.b64encode(self._build_errors_csv(lines)),
                "res_model": self._name,
                "res_id": self.id,
            }
        )
        return {
            "type": "ir.actions.act_url",
            "url": "/web/content/%s?download=true" % attachment.id,
            "target": "self",
        }

    @staticmethod
    def _build_errors_csv(lines):
        # ';' como delimitador y BOM utf-8: así lo abre bien Excel en
        # configuración regional argentina (que usa ',' como separador
        # decimal, por lo que toma ',' como parte de los números en vez de
        # como separador de columnas si el CSV usara ',').
        buffer = io.StringIO()
        writer = csv.writer(buffer, delimiter=";")
        writer.writerow(
            [
                _("Comprobante"),
                _("Código cliente"),
                _("Razón social"),
                _("Documento"),
                _("Fecha"),
                _("Importe total"),
                _("Error"),
            ]
        )
        for line in lines:
            writer.writerow(
                [
                    line.comprobante_display,
                    line.cliente_codigo,
                    line.cliente_razon_social,
                    line.cliente_documento,
                    line.fecha or "",
                    line.importe_total,
                    (line.error_message or "").replace("\n", " | "),
                ]
            )
        return buffer.getvalue().encode("utf-8-sig")

    def _reopen(self):
        return {
            "type": "ir.actions.act_window",
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
        }

    # ------------------------------------------------------------------
    # Lectura y parseo del Excel
    # ------------------------------------------------------------------

    @staticmethod
    def _normalize_header(value):
        return (value or "").strip().lower()

    @staticmethod
    def _to_str(value):
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    @staticmethod
    def _to_float(value):
        """Normaliza un valor numérico del Excel a float, tolerando que venga
        como número (int/float) o como texto con separador decimal ','
        (formato AR, ej. "1.234,56") o '.' (ej. "1234.56" o "1,234.56" con
        ',' de miles). Se usa para todas las columnas de importe/tasa
        (precio unitario, tasa de iva, tasa de iva no inscripto, importe iva
        inscripto/no inscripto, importe total neto del item, importe del
        renglón, montos y bases de impuestos internos/especiales, cantidad,
        importe total) para que no dependa de cómo haya venido formateada la
        celda en cada renglón."""
        if isinstance(value, (int, float)):
            return float(value)
        text = str(value).strip().replace(" ", "")
        if not text:
            raise ValueError("empty value")
        if "," in text and "." in text:
            if text.rindex(",") > text.rindex("."):
                text = text.replace(".", "").replace(",", ".")
            else:
                text = text.replace(",", "")
        elif "," in text:
            text = text.replace(",", ".")
        return float(text)

    @staticmethod
    def _digits_to_int(value):
        digits = "".join(ch for ch in str(value or "") if ch.isdigit())
        return int(digits) if digits else None

    @classmethod
    def _to_numeric_str(cls, value):
        """Normaliza a una representación canónica sin ceros a la izquierda,
        sea que la celda venga como texto, número o número con formato
        distinto entre renglones (ej. "0004" vs 4 vs 4.0). Se usa para
        pto_vta/numero_comprob: son parte de la clave de agrupación de la
        factura y también se comparan contra el diario, así que deben
        comparar igual sin importar cómo vino la celda."""
        digits = cls._digits_to_int(cls._to_str(value))
        return str(digits) if digits is not None else ""

    @classmethod
    def _to_cae(cls, value):
        text = cls._to_str(value)
        return text if text and text != "0" else ""

    @staticmethod
    def _to_date(value):
        text = str(int(value)) if isinstance(value, float) else str(value).strip()
        return datetime.strptime(text, "%Y%m%d").date()

    def _read_excel_rows(self, content):
        try:
            # read_only=True: lectura en modo streaming (no materializa todo
            # el workbook en un modelo de objetos mutable en memoria), mucho
            # más liviano para archivos de miles de filas. Compatible con el
            # único patrón de acceso usado acá (iter_rows(values_only=True)).
            workbook = openpyxl.load_workbook(
                io.BytesIO(content), read_only=True, data_only=True
            )
        except Exception as exc:  # pylint: disable=broad-except # noqa: BLE001
            raise UserError(
                _("No se pudo leer el archivo como Excel: %s") % exc
            ) from exc

        sheet = workbook.worksheets[0]
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            raise UserError(_("La primera hoja del archivo está vacía."))

        headers = [self._normalize_header(h) for h in header_row]
        missing_headers = [h for h in EXPECTED_HEADERS if h not in headers]
        if missing_headers:
            raise UserError(
                _("Faltan columnas en el archivo: %s") % ", ".join(missing_headers)
            )
        col_index = {h: i for i, h in enumerate(headers)}

        rows = []
        for row_number, raw_row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            if all(cell is None for cell in raw_row):
                continue
            rows.append(self._row_to_dict(raw_row, col_index, row_number))
        return rows

    def _row_to_dict(self, raw_row, col_index, row_number):
        def cell(header):
            idx = col_index[header]
            return raw_row[idx] if idx < len(raw_row) else None

        return {
            "_row_number": row_number,
            "tipo_comprobante": self._to_str(cell("tipo de comprobante")).upper(),
            "letra": self._to_str(cell("letra")).upper(),
            "pto_vta": self._to_numeric_str(cell("pto vta")),
            "numero_comprob": self._to_numeric_str(cell("numero comprob")),
            "fecha": cell("fecha"),
            "cod_cliente": self._to_str(cell("cod cliente")),
            "razon_social": self._to_str(cell("razon social")),
            "documento": self._to_str(cell("documento")),
            "fecha_vto": cell("fecha vto"),
            "importe_total": cell("importe total"),
            "comprobante_anulado": self._to_str(cell("comprobante anulado")).upper()
            == "S",
            "cae": self._to_cae(cell("cae")),
            "vto_cae": cell("vto cae"),
            "tipo_item": self._to_str(cell("tipo de item")),
            "cod_art": self._to_str(cell("cod art")),
            "cantidad": cell("cantidad"),
            "precio_unitario": cell("precio unitario"),
            "tasa_iva": cell("tasa de iva"),
            "tasa_iva_no_inscripto": cell("tasa de iva no inscripto"),
            "importe_iva_inscripto": cell("importe iva inscripto"),
            "importe_iva_no_inscripto": cell("importe iva no inscripto"),
            "importe_total_neto_item": cell("importe total neto del item"),
            "importe_del_renglon": cell("importe del renglon"),
            "special_taxes": [
                {
                    "cod": self._to_str(cell(cod_header)),
                    "monto": cell(monto_header),
                    "base": cell(base_header) if base_header else None,
                }
                for cod_header, monto_header, base_header in SPECIAL_TAX_COLUMNS
            ],
        }

    # ------------------------------------------------------------------
    # Cache de lookups para la previsualización: reemplaza los search()
    # repetidos por factura/línea/impuesto especial por un puñado de
    # búsquedas en lote (una por tipo de dato distinto en TODO el archivo),
    # más lookups en memoria. No cambia qué se resuelve, solo cómo.
    # ------------------------------------------------------------------

    def _build_preview_cache(self, groups):
        company = self.env.company
        cod_clientes, documentos, cod_arts, special_codes = (
            self._collect_preview_lookup_values(groups)
        )
        partner_by_bejerman, partner_by_vat = self._build_partner_caches(
            cod_clientes, documentos
        )
        product_by_bejerman, product_by_code = self._build_product_cache(cod_arts)
        perception_by_code, mapping_by_code = self._build_special_tax_caches(
            special_codes, company
        )
        sale_taxes = self.env["account.tax"].search(
            [("type_tax_use", "=", "sale"), ("company_id", "=", company.id)]
        )
        journal_cache = {}
        internal_number_groups, internal_number_existing = (
            self._build_internal_number_dedup_cache(groups, journal_cache, company)
        )

        return {
            "company": company,
            "voucher_type": {},
            "journal": journal_cache,
            "partner_by_bejerman": partner_by_bejerman,
            "partner_by_vat": partner_by_vat,
            "product_by_bejerman": product_by_bejerman,
            "product_by_code": product_by_code,
            "sale_taxes": sale_taxes,
            "perception_by_code": perception_by_code,
            "mapping_by_code": mapping_by_code,
            "moves_by_key": self._build_dedup_cache(
                groups, partner_by_bejerman, partner_by_vat
            ),
            "internal_number_groups": internal_number_groups,
            "internal_number_existing": internal_number_existing,
        }

    def _build_internal_number_dedup_cache(self, groups, journal_cache, company):
        """Prefetch para detectar colisiones contra la restricción real de
        Odoo account_move_unique_name_customers (número interno + diario +
        move_type + is_debit_note), que NO tiene en cuenta el partner. El
        dedup de _build_dedup_cache (por comprobante_ref + partner_id) no
        alcanza para detectar este caso: mismo número interno + diario ya
        usado por una factura de OTRO cliente, o repetido en otra fila del
        MISMO archivo (ver uso en _prepare_preview_line_vals).

        Usa/completa journal_cache de paso: la resolución de diario que hace
        acá queda memoizada y se reutiliza en el loop principal sin volver a
        buscarlo.

        La restricción de Odoo es un índice único PARCIAL: "UNIQUE (name,
        journal_id, move_type, is_debit_note) WHERE state = 'posted' AND
        name != '/' AND move_type IN (out_invoice, out_refund,
        out_receipt)". Por eso el chequeo contra facturas EXISTENTES filtra
        por state='posted': una factura ya cancelada con ese mismo número
        interno no bloquea nada en Odoo (comprobado: recrearla no viola la
        restricción). El chequeo dentro del MISMO archivo, en cambio, no se
        puede restringir así: dos filas del archivo que compartan la clave
        van a estar "posted" a la vez durante _confirm_chunk_fast (que
        postea todo el lote junto ANTES de cancelar las que corresponda),
        así que sí colisionan aunque después alguna termine cancelada.

        :return: tupla (groups_by_key, existing_move_by_key). groups_by_key
            mapea la clave a la lista de grupos que la comparten (para
            detectar duplicados dentro del archivo). existing_move_by_key
            mapea la clave a la factura YA existente y POSTED en Odoo que la
            usa (si hay alguna).
        """
        groups_by_key = {}
        for group in groups:
            if group.get("error"):
                continue
            doc_type = TIPO_COMPROBANTE_DOC_TYPES.get(group["tipo_comprobante"])
            if not doc_type:
                continue
            journal, _candidates = self._find_journal(
                group["letra"], group["pto_vta"], journal_cache, company
            )
            if not journal:
                continue
            move_type = (
                "out_refund" if group["tipo_comprobante"] == "NC" else "out_invoice"
            )
            is_debit_note = group["tipo_comprobante"] == "ND"
            internal_number = self._internal_number(
                group["pto_vta"], group["numero_comprob"]
            )
            key = (internal_number, journal.id, move_type, is_debit_note)
            groups_by_key.setdefault(key, []).append(group)

        existing_by_key = {}
        if groups_by_key:
            names = {k[0] for k in groups_by_key}
            journal_ids = {k[1] for k in groups_by_key}
            for move in self.env["account.move"].search(
                [
                    ("name", "in", list(names)),
                    ("journal_id", "in", list(journal_ids)),
                    ("state", "=", "posted"),
                ]
            ):
                mkey = (
                    move.name,
                    move.journal_id.id,
                    move.move_type,
                    move.is_debit_note,
                )
                if mkey in groups_by_key:
                    existing_by_key[mkey] = (
                        existing_by_key.get(mkey, self.env["account.move"]) | move
                    )
        return groups_by_key, existing_by_key

    def _collect_preview_lookup_values(self, groups):
        """Junta, en una sola pasada por todos los grupos, los valores
        DISTINTOS de todo el archivo que hacen falta para prefetchear
        cliente/producto/impuesto especial en lote (ver
        _build_preview_cache)."""
        cod_clientes = set()
        documentos = set()
        cod_arts = set()
        special_codes = set()
        for group in groups:
            if group.get("error"):
                continue
            if group["cod_cliente"]:
                cod_clientes.add(group["cod_cliente"])
            digits = self._only_digits(group["documento"])
            if digits:
                documentos.add(digits)
            for line in group["lines"]:
                if line["cod_art"]:
                    cod_arts.add(line["cod_art"])
                for slot in line["special_taxes"]:
                    if slot["cod"]:
                        special_codes.add(slot["cod"])
        return cod_clientes, documentos, cod_arts, special_codes

    def _build_partner_caches(self, cod_clientes, documentos):
        partner_by_bejerman = {}
        if cod_clientes:
            for partner in self.env["res.partner"].search(
                [("codigo_bejerman", "in", list(cod_clientes))]
            ):
                partner_by_bejerman[partner.codigo_bejerman] = (
                    partner_by_bejerman.get(
                        partner.codigo_bejerman, self.env["res.partner"]
                    )
                    | partner
                )

        partner_by_vat = {}
        if documentos:
            for partner in self.env["res.partner"].search(
                [("vat", "in", list(documentos))]
            ):
                partner_by_vat[partner.vat] = (
                    partner_by_vat.get(partner.vat, self.env["res.partner"]) | partner
                )
        return partner_by_bejerman, partner_by_vat

    def _build_product_cache(self, cod_arts):
        """Prefetch en lote para _find_product: por código Bejerman
        (product.template.codigo_bejerman, hoy prácticamente sin cargar en
        el catálogo) y, como respaldo, por código de artículo
        (product.product.default_code) — mismo criterio en cascada que
        _build_partner_caches usa con codigo_bejerman/vat."""
        product_by_bejerman = {}
        if cod_arts:
            for product in self.env["product.product"].search(
                [("codigo_bejerman", "in", list(cod_arts))]
            ):
                product_by_bejerman[product.codigo_bejerman] = (
                    product_by_bejerman.get(
                        product.codigo_bejerman, self.env["product.product"]
                    )
                    | product
                )

        product_by_code = {}
        if cod_arts:
            for product in self.env["product.product"].search(
                [("default_code", "in", list(cod_arts))]
            ):
                product_by_code[product.default_code] = (
                    product_by_code.get(
                        product.default_code, self.env["product.product"]
                    )
                    | product
                )
        return product_by_bejerman, product_by_code

    def _build_special_tax_caches(self, special_codes, company):
        perception_by_code = {}
        mapping_by_code = {}
        if special_codes:
            for tax in self.env["account.tax"].search(
                [
                    ("bejerman_code", "in", list(special_codes)),
                    ("tax_group_id.group_type", "=", "perception"),
                    ("company_id", "=", company.id),
                ]
            ):
                perception_by_code.setdefault(tax.bejerman_code, tax)
            for mapping in self.env["ivess.invoice.import.tax.code"].search(
                [("code", "in", list(special_codes)), ("company_id", "=", company.id)]
            ):
                mapping_by_code.setdefault(mapping.code, mapping.tax_id)
        return perception_by_code, mapping_by_code

    def _build_dedup_cache(self, groups, partner_by_bejerman, partner_by_vat):
        """Prefetch en lote del chequeo de deduplicación (ver uso en
        _prepare_preview_line_vals): junta la clave (move_type, ref,
        partner_id) de todas las facturas del archivo que sí resolvieron
        cliente, y hace una única búsqueda de account.move en vez de una
        por factura."""
        keys = set()
        for group in groups:
            if group.get("error"):
                continue
            partner = self._find_partner(
                group["cod_cliente"],
                group["documento"],
                partner_by_bejerman,
                partner_by_vat,
            )
            if not partner:
                continue
            move_type = (
                "out_refund" if group["tipo_comprobante"] == "NC" else "out_invoice"
            )
            ref = self._comprobante_ref(
                group["letra"], group["pto_vta"], group["numero_comprob"]
            )
            keys.add((move_type, ref, partner.id))

        moves_by_key = {}
        if not keys:
            return moves_by_key
        move_types = {k[0] for k in keys}
        refs = {k[1] for k in keys}
        partner_ids = {k[2] for k in keys}
        for move in self.env["account.move"].search(
            [
                ("move_type", "in", list(move_types)),
                ("ref", "in", list(refs)),
                ("partner_id", "in", list(partner_ids)),
            ]
        ):
            mkey = (move.move_type, move.ref, move.partner_id.id)
            moves_by_key[mkey] = moves_by_key.get(mkey, self.env["account.move"]) | move
        return moves_by_key

    # ------------------------------------------------------------------
    # Resolución de cada grupo contra Odoo (partner, tipo de comprobante,
    # productos, impuestos) y armado de los vals de la línea de
    # previsualización (se crean todas juntas en lote, ver action_preview).
    # ------------------------------------------------------------------

    def _prepare_preview_line_vals(self, group, cache):
        if group.get("error"):
            return {
                "wizard_id": self.id,
                "tipo_comprobante": group["tipo_comprobante"],
                "letra": group["letra"],
                "pto_vta": group["pto_vta"],
                "numero_comprobante": group["numero_comprob"],
                "has_error": True,
                "error_message": group["error"],
            }

        errors = []
        doc_type = TIPO_COMPROBANTE_DOC_TYPES.get(group["tipo_comprobante"])
        if not doc_type:
            errors.append(
                _(
                    "tipo de comprobante '%s' no soportado (esta versión solo"
                    " importa %s)."
                )
                % (
                    group["tipo_comprobante"],
                    ", ".join(sorted(TIPO_COMPROBANTE_DOC_TYPES)),
                )
            )
        move_type = "out_refund" if group["tipo_comprobante"] == "NC" else "out_invoice"
        is_debit_note = group["tipo_comprobante"] == "ND"

        voucher_type = self._find_voucher_type(
            group["letra"], doc_type, cache["voucher_type"]
        )
        if not voucher_type:
            errors.append(
                _("No se encontró un tipo de comprobante Odoo para la letra '%s'.")
                % group["letra"]
            )

        journal, journal_candidates = self._find_journal(
            group["letra"], group["pto_vta"], cache["journal"], cache["company"]
        )
        if not journal:
            if journal_candidates:
                errors.append(
                    _(
                        "La letra '%s' + punto de venta '%s' coincide con %s"
                        " diarios de venta distintos; debe coincidir con uno"
                        " solo."
                    )
                    % (group["letra"], group["pto_vta"], journal_candidates)
                )
            else:
                errors.append(
                    _(
                        "No se encontró un diario de venta para la letra '%s'"
                        " + punto de venta '%s'."
                    )
                    % (group["letra"], group["pto_vta"])
                )
        else:
            internal_number_error = self._check_internal_number_dedup(
                group, journal, move_type, is_debit_note, cache
            )
            if internal_number_error:
                errors.append(internal_number_error)

        partner = self._find_partner(
            group["cod_cliente"],
            group["documento"],
            cache["partner_by_bejerman"],
            cache["partner_by_vat"],
        )
        if not partner:
            errors.append(
                _(
                    "No se encontró un cliente con código Bejerman '%s' ni con"
                    " CUIT/documento '%s' (%s)."
                )
                % (group["cod_cliente"], group["documento"], group["razon_social"])
            )

        fecha, fecha_vto, vto_cae, importe_total, date_errors = (
            self._parse_preview_dates_and_amount(group)
        )
        errors.extend(date_errors)

        comprobante_ref = self._comprobante_ref(
            group["letra"], group["pto_vta"], group["numero_comprob"]
        )
        if partner:
            candidates = cache["moves_by_key"].get(
                (move_type, comprobante_ref, partner.id), self.env["account.move"]
            )
            if move_type == "out_invoice":
                # FC y ND comparten move_type "out_invoice": distinguirlos por
                # is_debit_note para no confundir una con la otra en el dedup.
                candidates = candidates.filtered(
                    lambda m: m.is_debit_note == is_debit_note
                )
            existing = candidates[:1]
            if existing:
                errors.append(
                    _(
                        "Ya existe una factura importada con esta clave (account.move #%s)."
                    )
                    % existing.id
                )

        detail_vals, detail_errors = self._resolve_detail_lines(group["lines"], cache)
        errors.extend(detail_errors)

        return {
            "wizard_id": self.id,
            "tipo_comprobante": group["tipo_comprobante"],
            "letra": group["letra"],
            "pto_vta": group["pto_vta"],
            "numero_comprobante": group["numero_comprob"],
            "comprobante_ref": comprobante_ref,
            "fecha": fecha,
            "fecha_vto": fecha_vto,
            "importe_total": importe_total,
            "comprobante_anulado": bool(group["comprobante_anulado"]),
            "cae": group["cae"],
            "vto_cae": vto_cae,
            "cliente_codigo": group["cod_cliente"],
            "cliente_razon_social": group["razon_social"],
            "cliente_documento": group["documento"],
            "partner_id": partner.id if partner else False,
            "voucher_type_id": voucher_type.id if voucher_type else False,
            "journal_id": journal.id if journal else False,
            "has_error": bool(errors),
            "error_message": "\n".join(errors) if errors else False,
            "detail_line_ids": detail_vals,
        }

    def _parse_preview_dates_and_amount(self, group):
        """Parsea fecha/fecha_vto/vto_cae/importe_total del grupo, cada uno
        de forma independiente (un valor inválido no impide parsear el
        resto). :return: tupla (fecha, fecha_vto, vto_cae, importe_total,
        errores)."""
        errors = []

        try:
            fecha = self._to_date(group["fecha"]) if group["fecha"] else False
        except ValueError:
            fecha = False
            errors.append(_("Fecha inválida: '%s'.") % group["fecha"])

        fecha_vto = False
        if group["fecha_vto"]:
            try:
                fecha_vto = self._to_date(group["fecha_vto"])
            except ValueError:
                errors.append(
                    _("Fecha de vencimiento inválida: '%s'.") % group["fecha_vto"]
                )

        vto_cae = False
        if group["vto_cae"]:
            try:
                vto_cae = self._to_date(group["vto_cae"])
            except ValueError:
                errors.append(
                    _("Vencimiento de CAE inválido: '%s'.") % group["vto_cae"]
                )

        try:
            importe_total = self._to_float(group["importe_total"])
        except ValueError:
            importe_total = 0.0
            errors.append(_("Importe total inválido: '%s'.") % group["importe_total"])

        return fecha, fecha_vto, vto_cae, importe_total, errors

    @staticmethod
    def _check_internal_number_dedup(group, journal, move_type, is_debit_note, cache):
        """Chequea el grupo contra la restricción real de Odoo
        account_move_unique_name_customers (número interno + diario + tipo
        de comprobante + is_debit_note), que NO tiene en cuenta el cliente:
        a diferencia del dedup por comprobante_ref + partner_id (que solo
        detecta "ya importé ESTA factura para ESTE cliente"), acá puede
        colisionar tanto con una factura ya existente de OTRO cliente como
        con otra fila del MISMO archivo que resuelve al mismo diario (dato
        real encontrado: el preview daba OK a ambas filas, y recién al
        confirmar fallaba con un IntegrityError de Postgres).

        :return: mensaje de error (str) si hay colisión, o None."""
        internal_number = IvessInvoiceImportWizard._internal_number(
            group["pto_vta"], group["numero_comprob"]
        )
        key = (internal_number, journal.id, move_type, is_debit_note)
        existing_move = cache["internal_number_existing"].get(key)
        if existing_move:
            return _(
                "El número interno '%s' del diario '%s' ya está usado por"
                " otra factura CONFIRMADA en Odoo (account.move #%s, cliente"
                " '%s'); Odoo no permite crear otra con esa misma"
                " combinación mientras esa siga posteada, más allá del"
                " cliente."
            ) % (
                internal_number,
                journal.display_name,
                existing_move[:1].id,
                existing_move[:1].partner_id.display_name,
            )
        other_rows = len(cache["internal_number_groups"].get(key, [])) - 1
        if other_rows > 0:
            return _(
                "El número interno '%s' del diario '%s' se repite en otra(s)"
                " %s fila(s) del mismo archivo; Odoo exige que sea único por"
                " diario, más allá del cliente."
            ) % (internal_number, journal.display_name, other_rows)
        return None

    @staticmethod
    def _comprobante_ref(letra, pto_vta, numero_comprob):
        """Clave de deduplicación/referencia externa del comprobante, en el
        mismo formato compuesto que usa el propio sistema origen (aguas) para
        identificar un comprobante (ver hoja "items" de sus exportes: letra +
        punto de venta + número). Se guarda en el campo nativo
        account.move.ref."""
        return "%s%s%s" % (
            letra,
            (pto_vta or "").zfill(4),
            (numero_comprob or "").zfill(8),
        )

    def _find_voucher_type(self, letra, doc_type, voucher_type_cache):
        if not letra or not doc_type:
            return None
        key = (letra, doc_type)
        if key not in voucher_type_cache:
            candidates = self.env["res.voucher.type"].search(
                [("doc_type", "=", doc_type), ("denomination", "=", letra.lower())]
            )
            # Puede haber más de un tipo para la misma letra (ej: "FACTURAS
            # A", "... CON LEYENDA RETENCIÓN", "... MiPyMEs FCE"). Sin una
            # columna en el Excel que distinga el caso especial, se asume el
            # comprobante "plano" (menor afip_code) como default razonable
            # para una importación masiva estándar.
            voucher_type_cache[key] = (
                candidates.sorted(key=lambda v: v.afip_code)[0] if candidates else None
            )
        return voucher_type_cache[key]

    def _find_journal(self, letra, pto_vta, journal_cache, company):
        """Resuelve el diario de ventas a partir de la letra + punto de venta
        del Excel (columnas "letra" y "pto vta"): el archivo mezcla
        comprobantes de más de un diario, así que el diario ya no se elige a
        mano en el wizard. Un mismo punto de venta AFIP puede tener un
        diario Odoo distinto por letra (ej. "FACTURAS A 0001" vs "FACTURAS B
        0001"), así que primero se filtra por account.journal.denomination
        (igual que res.voucher_type.denomination en _find_voucher_type) y
        recién ahí se matchea el número: contra el código AFIP del diario
        (account.journal.code, ver get_pos_number() en l10n_ar_eynes) o, si
        no coincide con ninguno, contra los dígitos del nombre del diario
        (algunos diarios están identificados solo por su nombre: el campo
        "code" puede no tener nada que ver con el punto de venta real, ej.
        un diario "9999" con code="INV1"). NO existe account.journal.
        l10n_ar_afip_pos_number en esta instalación: el módulo Odoo
        oficial "l10n_ar" está desinstalado, solo corre "l10n_ar_eynes"
        (fork propio), que no tiene ese campo.

        El resultado se memoiza por (letra, pto_vta) en journal_cache: la
        misma combinación se repite en muchas facturas del archivo y no
        cambia dentro de una misma previsualización.

        :return: tupla (diario o None, cantidad de diarios candidatos).
        """
        pos_number = self._digits_to_int(pto_vta)
        if not letra or pos_number is None:
            return None, 0
        key = (letra, pto_vta)
        if key not in journal_cache:
            journals = self.env["account.journal"].search(
                [
                    ("type", "=", "sale"),
                    ("company_id", "=", company.id),
                    ("denomination", "=", letra.lower()),
                ]
            )
            matches = journals.filtered(
                lambda j: j.get_pos_number() == pos_number
                or self._digits_to_int(j.name) == pos_number
            )
            journal_cache[key] = (
                matches[0] if len(matches) == 1 else None,
                len(matches),
            )
        return journal_cache[key]

    @staticmethod
    def _only_digits(value):
        return "".join(ch for ch in (value or "") if ch.isdigit())

    def _find_partner(
        self, cod_cliente, documento, partner_by_bejerman, partner_by_vat
    ):
        """Resuelve el cliente por su código Bejerman (columna "cod cliente"
        del Excel, matcheado contra res.partner.codigo_bejerman) y, si no
        matchea, por CUIT/documento (res.partner.vat) como respaldo.
        partner_by_bejerman/partner_by_vat vienen prefetcheados en lote (ver
        _build_preview_cache) para todos los valores distintos del archivo."""
        if cod_cliente:
            candidates = partner_by_bejerman.get(cod_cliente, self.env["res.partner"])
            if len(candidates) == 1:
                return candidates[0]

        digits = self._only_digits(documento)
        if not digits:
            return None
        candidates = partner_by_vat.get(digits, self.env["res.partner"])
        return candidates[0] if len(candidates) == 1 else None

    @staticmethod
    def _find_tax(tasa_iva, sale_taxes):
        candidates = sale_taxes.filtered(
            lambda t: round(t.amount, 2) == round(tasa_iva, 2)
        )
        return candidates[0] if len(candidates) == 1 else None, len(candidates)

    def _resolve_detail_lines(self, lines, cache):
        detail_vals = []
        errors = []

        if not lines:
            errors.append(_("La factura no tiene líneas de detalle."))
            return detail_vals, errors

        for line in lines:
            line_errors = []

            try:
                cantidad = self._to_float(line["cantidad"])
            except (TypeError, ValueError):
                cantidad = 0.0
                line_errors.append(_("Cantidad inválida: '%s'.") % line["cantidad"])

            try:
                precio_unitario = self._to_float(line["precio_unitario"])
            except (TypeError, ValueError):
                precio_unitario = 0.0
                line_errors.append(
                    _("Precio unitario inválido: '%s'.") % line["precio_unitario"]
                )

            try:
                tasa_iva = self._to_float(line["tasa_iva"])
            except (TypeError, ValueError):
                tasa_iva = 0.0
                line_errors.append(_("Tasa de IVA inválida: '%s'.") % line["tasa_iva"])

            try:
                tasa_iva_no_inscripto = self._to_float(line["tasa_iva_no_inscripto"])
            except (TypeError, ValueError):
                tasa_iva_no_inscripto = 0.0
            if tasa_iva_no_inscripto:
                line_errors.append(
                    _("Tasa de IVA no inscripto (%.2f) no soportada en esta versión.")
                    % tasa_iva_no_inscripto
                )

            product = self._find_product(
                line["cod_art"], cache["product_by_bejerman"], cache["product_by_code"]
            )
            if not product:
                line_errors.append(
                    _("No se encontró un producto con código de artículo '%s'.")
                    % line["cod_art"]
                )

            # Tasa de IVA 0: la línea no lleva impuesto de IVA (tax_id queda
            # vacío), no se busca ni exige que exista una tasa "0%" en Odoo.
            tax = None
            if not line_errors and tasa_iva:
                tax, tax_candidates = self._find_tax(tasa_iva, cache["sale_taxes"])
                if not tax:
                    line_errors.append(
                        _(
                            "No se encontró (o es ambigua, %s candidatos) la tasa de IVA %.2f%%."
                        )
                        % (tax_candidates, tasa_iva)
                    )

            try:
                importe_total_neto_item = self._to_float(
                    line["importe_total_neto_item"]
                )
            except (TypeError, ValueError):
                importe_total_neto_item = 0.0

            try:
                importe_iva_inscripto = self._to_float(line["importe_iva_inscripto"])
            except (TypeError, ValueError):
                importe_iva_inscripto = 0.0

            try:
                importe_iva_no_inscripto = self._to_float(
                    line["importe_iva_no_inscripto"]
                )
            except (TypeError, ValueError):
                importe_iva_no_inscripto = 0.0

            special_tax_vals, special_tax_errors = self._resolve_special_taxes(
                line["special_taxes"],
                cache["perception_by_code"],
                cache["mapping_by_code"],
            )
            line_errors.extend(special_tax_errors)

            try:
                importe_del_renglon = self._to_float(line["importe_del_renglon"])
            except (TypeError, ValueError):
                importe_del_renglon = 0.0

            detail_vals.append(
                (
                    0,
                    0,
                    {
                        "tipo_item": line["tipo_item"],
                        "cod_art": line["cod_art"],
                        "product_id": product.id if product else False,
                        "cantidad": cantidad,
                        "precio_unitario": precio_unitario,
                        "tasa_iva": tasa_iva,
                        "tax_id": tax.id if tax else False,
                        "importe_total_neto_item": importe_total_neto_item,
                        "importe_iva_inscripto": importe_iva_inscripto,
                        "importe_iva_no_inscripto": importe_iva_no_inscripto,
                        "importe_del_renglon": importe_del_renglon,
                        "special_tax_ids": special_tax_vals,
                        "has_error": bool(line_errors),
                        "error_message": "; ".join(line_errors)
                        if line_errors
                        else False,
                    },
                )
            )
            errors.extend(line_errors)

        return detail_vals, errors

    def _resolve_special_taxes(
        self, special_taxes, perception_by_code, mapping_by_code
    ):
        """Resuelve las columnas de impuesto especial/interno de una línea
        (ver SPECIAL_TAX_COLUMNS): cada una matchea por separado contra un
        impuesto Odoo (ver _find_special_tax) y aporta su propio monto y
        base imponible. "cod impuesto interno" no trae columna de base en el
        Excel y el impuesto interno no tiene base imponible: su base se
        informa siempre en 0. A diferencia del resto de los impuestos
        especiales, si el impuesto interno viene con monto en 0 o vacío no
        es un error: significa que esa línea no tiene impuesto interno y se
        ignora esa columna sin frenar la importación."""
        special_tax_vals = []
        errors = []

        for slot in special_taxes:
            cod = slot["cod"]
            if not cod:
                continue

            try:
                monto = self._to_float(slot["monto"])
            except (TypeError, ValueError):
                monto = 0.0

            has_base_column = slot["base"] is not None

            if not has_base_column and not monto:
                continue

            base = 0.0
            if has_base_column:
                try:
                    base = self._to_float(slot["base"])
                except (TypeError, ValueError):
                    base = 0.0

            special_tax = self._find_special_tax(
                cod, perception_by_code, mapping_by_code
            )
            if not special_tax:
                errors.append(
                    _(
                        "No se encontró un mapeo para el código de impuesto"
                        " especial '%s' (configuralo en Contabilidad >"
                        " Configuración > Códigos de impuesto especial"
                        " (importación))."
                    )
                    % cod
                )
                continue
            if not monto or (has_base_column and not base):
                errors.append(
                    _(
                        "El código de impuesto especial '%s' está mapeado a"
                        " '%s' pero el monto%s informado es 0."
                    )
                    % (
                        cod,
                        special_tax.name,
                        _(" o la base") if has_base_column else "",
                    )
                )
                continue

            special_tax_vals.append(
                (
                    0,
                    0,
                    {
                        "cod": cod,
                        "tax_id": special_tax.id,
                        "monto": monto,
                        "base": base,
                    },
                )
            )

        return special_tax_vals, errors

    @staticmethod
    def _find_special_tax(cod_impuesto_especial, perception_by_code, mapping_by_code):
        """Resuelve un código de impuesto especial/interno contra los dos
        prefetch de _build_preview_cache:
        1) Percepciones: matcheadas por código Bejerman cargado en el propio
           impuesto (Contabilidad > Impuestos > pestaña "Perceptions").
        2) Impuestos internos (u otras percepciones sin código Bejerman
           cargado todavía): mapeo manual configurable
           (ivess.invoice.import.tax.code)."""
        perception = perception_by_code.get(cod_impuesto_especial)
        if perception:
            return perception
        return mapping_by_code.get(cod_impuesto_especial)

    def _find_product(self, cod_art, product_by_bejerman, product_by_code):
        """Resuelve el producto por código Bejerman
        (product.template.codigo_bejerman) y, si no matchea, por código de
        artículo (product.product.default_code) como respaldo — mismo
        criterio en cascada que _find_partner con codigo_bejerman/vat."""
        if not cod_art:
            return None
        candidates = product_by_bejerman.get(cod_art, self.env["product.product"])
        if len(candidates) == 1:
            return candidates[0]
        candidates = product_by_code.get(cod_art, self.env["product.product"])
        return candidates[0] if len(candidates) == 1 else None

    # ------------------------------------------------------------------
    # Paso 2 -> 3: crear las facturas que no tengan error
    # ------------------------------------------------------------------

    # Cantidad de facturas por tanda en el camino rápido de action_confirm
    # (ver _confirm_chunk_fast): bastante más grande que "unas pocas" para
    # amortizar el overhead fijo de create()/action_post() (medido: ~4x
    # menos queries por factura al crear 20 juntas que una por una). Ya no
    # hace falta que sea chica "por las dudas" de que falle: si una tanda
    # falla, se bisecta en vez de caer directo a fila por fila (ver
    # _confirm_chunk), así que una tanda grande con un par de filas
    # problemáticas sigue aprovechando el batch para el resto.
    _CONFIRM_CHUNK_SIZE = 500
    # Por debajo de este tamaño, ante una falla se cae directo al camino
    # serie en vez de seguir bisectando (procesar ~10 filas una por una ya
    # es rápido; seguir partiendo no ahorraría nada).
    _CONFIRM_MIN_BISECT_SIZE = 10

    def action_confirm(self):
        self.ensure_one()
        # tracking_disable evita que cada create()/write() de account.move
        # disloque el mecanismo de chatter de mail.thread (seguidores,
        # tracking de cambios, mensaje de "creado"): no aporta nada en una
        # importación histórica y es una porción medible del tiempo por
        # factura (medido con cProfile contra un archivo real).
        self = self.with_context(tracking_disable=True)

        for result_line in self.result_line_ids.filtered("has_error"):
            result_line.resultado = "error"

        pending_lines = self.result_line_ids.filtered(lambda r: not r.has_error)
        for start in range(0, len(pending_lines), self._CONFIRM_CHUNK_SIZE):
            chunk = pending_lines[start : start + self._CONFIRM_CHUNK_SIZE]
            self._confirm_chunk(chunk)
            # Se confirma cada tanda por separado en vez de dejar TODO el
            # import en una única transacción: si el proceso se corta a
            # mitad de camino (timeout del navegador/proxy, reinicio), lo
            # que ya se procesó queda guardado en vez de perderse entero
            # (encontrado en un import real: la transacción completa tardó
            # >10 minutos sin que nada quedara visible hasta el final).
            # pylint: disable=invalid-commit
            self.env.cr.commit()

        self.ok_count = len(
            self.result_line_ids.filtered(lambda r: r.resultado == "ok")
        )
        self.error_count = len(
            self.result_line_ids.filtered(lambda r: r.resultado == "error")
        )
        self.state = "done"
        return self._reopen()

    def _confirm_chunk(self, chunk):
        """Intenta el camino rápido (_confirm_chunk_fast) para la tanda
        completa; si falla, la parte a la mitad y reintenta cada mitad por
        separado (recursivo), en vez de caer directo a fila por fila.

        Motivo (encontrado en un import real): con una tanda grande
        "todo o nada", bastaba con que 1 de cada ~50 filas tuviera un
        problema imprevisible (ej. dos res.partner con el mismo CUIT) para
        que estadísticamente CASI TODAS las tandas tuvieran al menos una
        falla y cayeran enteras al modo serie, perdiendo casi toda la
        ganancia del batch. Bisectando, solo la sub-tanda que realmente
        contiene la fila problemática paga el costo de ir más despacio; el
        resto de la tanda se resuelve en lote de todos modos.
        """
        if not chunk:
            return
        if len(chunk) <= self._CONFIRM_MIN_BISECT_SIZE:
            if not self._confirm_chunk_fast(chunk):
                self._confirm_chunk_serial(chunk)
            return
        if self._confirm_chunk_fast(chunk):
            return
        mid = len(chunk) // 2
        self._confirm_chunk(chunk[:mid])
        self._confirm_chunk(chunk[mid:])

    def _confirm_chunk_fast(self, chunk):
        """Crea y postea TODAS las facturas de la tanda de una sola vez (un
        create() y un action_post() en lote), en vez de una por una: mucho
        más rápido cuando ninguna de la tanda tiene problemas (medido:
        create() en lote reduce las queries por factura de ~98 a ~26). Si
        CUALQUIER factura de la tanda falla (create, post o cancelación), se
        descarta la tanda entera (rollback del savepoint) y no se toca
        ningún result_line: el llamador cae a _confirm_chunk_serial, que
        procesa fila por fila con el mismo aislamiento de errores que la
        versión anterior a esta optimización.

        :return: True si la tanda quedó resuelta (todas OK); False si hubo
            que abortarla y no se tocó nada.
        """
        try:
            with self.env.cr.savepoint():
                move_vals_list = [
                    self._move_vals(
                        result_line, self._invoice_wide_special_tax_ids(result_line)
                    )
                    for result_line in chunk
                ]
                moves = self.env["account.move"].create(move_vals_list)
                for result_line, move in zip(chunk, moves, strict=True):
                    self._set_move_line_taxes(
                        move,
                        result_line,
                        self._invoice_wide_special_tax_ids(result_line),
                    )
                moves.action_post()
                to_cancel = self.env["account.move"]
                for result_line, move in zip(chunk, moves, strict=True):
                    if result_line.comprobante_anulado:
                        to_cancel |= move
                if to_cancel:
                    to_cancel.button_cancel()
                for result_line, move in zip(chunk, moves, strict=True):
                    result_line.write({"resultado": "ok", "odoo_move_id": move.id})
            return True
        except Exception:  # pylint: disable=broad-except # noqa: BLE001
            return False

    def _confirm_chunk_serial(self, chunk):
        """Camino de siempre: una factura a la vez, cada una con su propio
        savepoint, para que un error puntual (ej. una restricción de Odoo
        que no se puede prever en el preview) no se lleve puestas al resto
        de la tanda."""
        for result_line in chunk:
            try:
                with self.env.cr.savepoint():
                    move = self._create_move(result_line)
                    move.action_post()
                    if result_line.comprobante_anulado:
                        move.button_cancel()
                result_line.write({"resultado": "ok", "odoo_move_id": move.id})
            except Exception as exc:  # pylint: disable=broad-except # noqa: BLE001
                result_line.write(
                    {
                        "resultado": "error",
                        "has_error": True,
                        "error_message": _("Error al crear la factura: %s") % exc,
                    }
                )

    def _create_move(self, result_line):
        invoice_wide_tax_ids = self._invoice_wide_special_tax_ids(result_line)
        move_vals = self._move_vals(result_line, invoice_wide_tax_ids)
        move = self.env["account.move"].create(move_vals)
        self._set_move_line_taxes(move, result_line, invoice_wide_tax_ids)
        return move

    def _move_vals(self, result_line, invoice_wide_tax_ids):
        line_vals = [
            (
                0,
                0,
                {
                    "product_id": detail.product_id.id,
                    "name": detail.product_id.display_name,
                    "quantity": detail.cantidad,
                    "price_unit": detail.precio_unitario,
                    "tax_ids": [
                        (6, 0, self._line_tax_ids(detail, invoice_wide_tax_ids))
                    ],
                },
            )
            for detail in result_line.detail_line_ids
        ]
        move_vals = {
            "move_type": "out_refund"
            if result_line.tipo_comprobante == "NC"
            else "out_invoice",
            "journal_id": result_line.journal_id.id,
            "partner_id": result_line.partner_id.id,
            "voucher_type_id": result_line.voucher_type_id.id,
            "invoice_date": result_line.fecha,
            "invoice_line_ids": line_vals,
            "ref": result_line.comprobante_ref,
            "cae": result_line.cae or False,
            "cae_due_date": result_line.vto_cae or False,
            # Esta es una factura histórica, ya emitida y numerada en el
            # sistema origen (con CAE propio): se fija el número real acá en
            # vez de dejar que se autonumere con la secuencia del diario.
            # posted_before=True es lo que l10n_ar_eynes chequea en
            # account.move._post() para no pisar internal_number con el
            # próximo valor de la secuencia (ver l10n_ar_eynes/models/
            # account_move.py).
            "internal_number": self._internal_number(
                result_line.pto_vta, result_line.numero_comprobante
            ),
            "posted_before": True,
            # Las percepciones/impuestos internos de esta factura ya vienen
            # calculados del sistema origen (columnas "monto imp interno" /
            # "monto imp especiales*"): no queremos que l10n_ar_eynes intente
            # recalcularlos solo con la posición fiscal del cliente.
            "disable_perceptions": True,
        }
        if result_line.tipo_comprobante == "ND":
            move_vals["is_debit_note"] = True
        if result_line.fecha_vto:
            move_vals["invoice_date_due"] = result_line.fecha_vto
        perception_vals, internal_tax_vals = self._special_tax_vals(result_line)
        if perception_vals:
            move_vals["perception_ids"] = perception_vals
        if internal_tax_vals:
            move_vals["internal_taxes_ids"] = internal_tax_vals
        return move_vals

    def _set_move_line_taxes(self, move, result_line, invoice_wide_tax_ids):
        # account.move.line.tax_ids es un compute (store=True) que se
        # recalcula por el solo hecho de tener product_id seteado (ver
        # AccountMoveLine._get_computed_taxes() en l10n_ar_eynes: agrega ahí
        # los impuestos internos por defecto de la posición fiscal del
        # cliente a CUALQUIER línea de producto, sin mirar detail_line_ids).
        # Para que cada línea quede EXACTAMENTE con los impuestos que
        # matcheamos para ella arriba (ni de más por ese compute, ni de
        # menos), se reescribe tax_ids explícito después del create(): un
        # write() de tax_ids solo no depende de product_id/product_uom_id,
        # así que no dispara ese compute de nuevo.
        for line, detail in zip(
            move.invoice_line_ids.filtered(lambda l: l.display_type == "product"),
            result_line.detail_line_ids,
            strict=False,
        ):
            line.tax_ids = [(6, 0, self._line_tax_ids(detail, invoice_wide_tax_ids))]
        # perception_ids/internal_taxes_ids por sí solos solo alimentan las
        # pestañas "Percepciones"/"Internal taxes": el importe informado ahí
        # también tiene que quedar como monto de la línea de impuesto real
        # del asiento (generada por Odoo a partir del tax_ids de cada línea,
        # reescrito arriba). NO se usan acá los helpers
        # link_perception_to_move_lines()/link_internal_taxes_to_move_lines()
        # de l10n_ar_eynes: con disable_perceptions=True esos recalculan la
        # aplicabilidad de percepciones por posición fiscal/cuenta contable
        # y las aplican a TODAS las líneas de producto sin distinguir
        # percepción de IVA (por línea) de percepción IIBB (toda la
        # factura); acá se arma explícito el tax_ids correcto por línea
        # (ver _line_tax_ids) tal como matcheamos arriba.
        perception_vals, internal_tax_vals = self._special_tax_vals(result_line)
        if perception_vals:
            move._update_perception_move_line_amount()
        if internal_tax_vals:
            move._update_internal_taxes_move_line_amount()

    @staticmethod
    def _invoice_wide_special_tax_ids(result_line):
        """IDs de los impuestos especiales que aplican a TODA la factura y
        no solo a la línea del Excel donde se detectaron: las percepciones
        de IIBB (ARBA/AGIP-CABA, account.tax.retention_type ==
        "gross_income"). A diferencia del impuesto interno y la percepción
        de IVA (que sí son por línea de producto), IIBB se calcula sobre el
        neto de toda la factura, así que su tax_id debe quedar en el
        tax_ids de todas las líneas de producto del asiento (ver
        _line_tax_ids) para que Odoo calcule bien la base de esa línea de
        impuesto (tax_base_amount / tax_totals)."""
        special_taxes = result_line.detail_line_ids.mapped("special_tax_ids")
        return special_taxes.filtered(
            lambda s: s.tax_id.retention_type == "gross_income"
        ).tax_id.ids

    @staticmethod
    def _line_tax_ids(detail, invoice_wide_tax_ids):
        """IDs de impuesto a asignar a una línea de producto del asiento: su
        IVA (si tiene: con tasa de IVA 0 la línea no lleva tax_id de IVA),
        sus impuestos especiales/internos propios que NO sean IIBB (impuesto
        interno y percepción de IVA: van por línea, ver
        _resolve_special_taxes) y las percepciones IIBB de toda la factura
        (invoice_wide_tax_ids, ver _invoice_wide_special_tax_ids)."""
        per_line_tax_ids = detail.special_tax_ids.filtered(
            lambda s: s.tax_id.retention_type != "gross_income"
        ).tax_id.ids
        iva_tax_ids = [detail.tax_id.id] if detail.tax_id else []
        return list(
            dict.fromkeys(iva_tax_ids + per_line_tax_ids + invoice_wide_tax_ids)
        )

    @staticmethod
    def _internal_number(pto_vta, numero_comprobante):
        # Se arma con el "pto vta" tal como vino del Excel (ya validado
        # contra el diario en _find_journal), NO releyendo el diario: los
        # campos de punto de venta del diario (account.journal.code) pueden
        # no tener relación con el punto de venta real (ver comentario en
        # _find_journal, ej. diario "9999" con code="INV1"). Mismo padding
        # que _comprobante_ref/comprobante_display (zfill(4) en el punto de
        # venta), ya que replica el formato compuesto del sistema origen.
        return "%s-%s" % (
            (pto_vta or "").zfill(4),
            (numero_comprobante or "").zfill(8),
        )

    @staticmethod
    def _special_tax_vals(result_line):
        """Agrega, por impuesto especial resuelto (percepción IIBB/IVA,
        impuesto interno), una línea a perception_ids o internal_taxes_ids
        con base y monto sumados de todas las líneas de detalle que
        comparten ese impuesto (no se recalculan: se toman tal cual del
        archivo origen)."""
        perception_vals = []
        internal_tax_vals = []
        special_lines = result_line.detail_line_ids.mapped("special_tax_ids")
        for tax in special_lines.mapped("tax_id"):
            lines = special_lines.filtered(lambda s, tax=tax: s.tax_id == tax)
            base = sum(lines.mapped("base"))
            amount = sum(lines.mapped("monto"))
            if tax.tax_group_id.group_type == "internals":
                internal_tax_vals.append(
                    (
                        0,
                        0,
                        {
                            # "name" es un campo compute/store, pero el compute no
                            # se dispara de forma confiable al crear vía comandos
                            # (0,0,{...}) anidados en account.move.create(): se
                            # informa explícito, igual que hace l10n_ar_eynes.
                            "name": tax.name,
                            "tax_id": tax.id,
                            "base": base,
                            "amount": amount,
                        },
                    )
                )
            else:
                perception_vals.append(
                    (
                        0,
                        0,
                        {
                            "name": tax.name,
                            "perception_id": tax.id,
                            "partner_id": result_line.partner_id.id,
                            "base": base,
                            "amount": amount,
                        },
                    )
                )
        return perception_vals, internal_tax_vals
