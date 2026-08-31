import base64
import io
import json
import re
import unicodedata
import zipfile
from collections import Counter
from datetime import datetime

from markupsafe import escape

from odoo import _, fields, models
from odoo.exceptions import UserError

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None

# openpyxl no puede leer (ni en modo read_only) una hoja cuyo <autoFilter>
# tiene un <customFilter operator="..."> con un valor que no reconoce
# (pasa con archivos tocados por ciertas versiones de Excel/Sheets/
# LibreOffice). El filtro es solo metadata de UI de la planilla, no afecta
# los datos, así que se lo saca del XML antes de abrir el archivo.
_AUTOFILTER_RE = re.compile(
    rb"<autoFilter\b(?:[^>]*?/>|[^>]*?>.*?</autoFilter>)", re.DOTALL
)

# Columna del Excel (normalizada: sin acentos, minúscula) -> clave canónica
# interna. "_ignored" son columnas confirmadas 100% vacías en el maestro real
# (nrosub, Vendedor, Empresa, Cuenta por cobrar, Número de registro de IIBB):
# se exigen en el archivo para detectar un formato de exportación distinto,
# pero no se usan para nada. "Etiqueta"/"Etiquetas" varía entre las dos hojas
# del archivo real (singular en una, plural en la otra), de ahí el alias.
HEADER_MAP = {
    "codigo de cliente": "customer_code",
    "codigo bejerman": "codigo_bejerman",
    "nrosub": "_ignored",
    "tipo de empresa": "tipo_empresa",
    "nombre": "nombre",
    "correo electronico": "email",
    "telefono": "phone",
    "numero de celular": "mobile",
    "observaciones de direccion": "obs_direccion",
    "calle": "calle",
    "numero": "numero",
    "pais": "pais",
    "ciudad": "ciudad",
    "estado": "estado",
    "codigo postal": "zip",
    "tipo de documento": "tipo_documento",
    "nrcuit": "vat",
    "numero de registro de iibb": "_ignored",
    "cliente importante": "cliente_importante",
    "fecha de alta": "fecha_alta",
    "vendedor": "_ignored",
    "precios especiales": "precios_especiales",
    "posicion fiscal": "posicion_fiscal",
    "empresa": "_ignored",
    "tipo de cliente": "tipo_cliente",
    "cuenta por cobrar": "_ignored",
    "requiere comprobante": "requiere_comprobante",
    "geo latitud": "lat",
    "geo longitud": "lon",
    "notas": "notas",
    "etiqueta": "etiqueta",
    "etiquetas": "etiqueta",
}
REQUIRED_KEYS = sorted({v for v in HEADER_MAP.values() if v != "_ignored"})

# "Estado" en el Excel viene como display_name armado a mano en el sistema
# origen ("Buenos Aires (AR)"), no como el name real de res.country.state:
# hay que pelar el sufijo de país antes de buscar.
STATE_COUNTRY_SUFFIX_RE = re.compile(r"\s*\([a-z]{2}\)$")

RES_PARTNER_COLUMNS = [
    "name",
    "is_company",
    "email",
    "phone",
    "mobile_number",
    "street",
    "street_name",
    "street_number",
    "city",
    "zip",
    "country_id",
    "state_id",
    "vat",
    "document_type_id",
    "property_account_position_id",
    "partner_type_id",
    "customer_code",
    "codigo_bejerman",
    "fecha_alta",
    "is_important_client",
    "requiere_comprobante",
    "has_special_price",
    "comment",
    "partner_latitude",
    "partner_longitude",
    "create_date",
    "write_date",
    "create_uid",
    "write_uid",
    "company_id",
    "active",
    "type",
    # required=True con default en Odoo core/purchase_stock: el INSERT
    # crudo no aplica defaults, y sin esto Odoo rechaza el primer write()
    # posterior (aunque sea de otro campo) por validación de campo
    # requerido en todo el registro.
    "autopost_bills",
    "group_rfq",
    "group_on",
]


def _normalize(value):
    """minúscula, sin espacios extra, sin acentos -- usado tanto para
    encabezados del Excel como para los valores que se matchean contra
    nombres de registros existentes (país, estado, tipo de documento, etc.)."""
    text = (value or "").strip().lower()
    return "".join(
        ch
        for ch in unicodedata.normalize("NFKD", text)
        if not unicodedata.combining(ch)
    )


class ResPartnerImportWizard(models.TransientModel):
    _name = "res.partner.import.wizard"
    _description = "Importador de clientes (maestro) desde archivo Excel"

    file = fields.Binary(string="Archivo (.xlsx)", required=True)
    filename = fields.Char(string="Nombre de archivo")
    row_limit = fields.Integer(
        string="Límite de filas (opcional, para pruebas)",
        help="Si se completa, solo se procesan las primeras N filas del"
        " Excel (contando ambas hojas en orden). Dejar vacío para procesar"
        " el archivo completo.",
    )
    state = fields.Selection(
        [
            ("upload", "Subir archivo"),
            ("preview", "Análisis previo"),
            ("done", "Resultado"),
        ],
        default="upload",
    )
    issue_ids = fields.One2many(
        "res.partner.import.wizard.issue", "wizard_id", string="Observaciones"
    )
    total_rows = fields.Integer(string="Filas leídas", readonly=True)
    rows_to_insert = fields.Integer(string="Filas a importar", readonly=True)
    rows_skipped = fields.Integer(string="Filas excluidas", readonly=True)
    geo_ok_count = fields.Integer(string="Con coordenadas", readonly=True)
    geo_blank_count = fields.Integer(string="Sin coordenadas", readonly=True)
    inserted_count = fields.Integer(string="Clientes creados", readonly=True)

    # ------------------------------------------------------------------
    # Paso 1 -> 2: leer y resolver todo en memoria, sin tocar res.partner.
    # ------------------------------------------------------------------

    def action_preview(self):
        self.ensure_one()
        if not self.file:
            raise UserError(_("Adjuntá un archivo para importar."))
        if openpyxl is None:
            raise UserError(_("Falta la librería 'openpyxl' en el servidor (.xlsx)"))

        raw_rows = self._parse_workbook(base64.b64decode(self.file))
        lookups = self._build_lookups()
        bejerman_counts = self._count_bejerman_codes(raw_rows)

        issue_counter = Counter()
        skipped = 0
        geo_ok = 0
        geo_blank = 0

        for raw in raw_rows:
            resolved = self._resolve_row(raw, lookups, bejerman_counts)
            for category, value in resolved["issues"]:
                issue_counter[(category, value)] += 1
            if resolved["skip"]:
                skipped += 1
            elif resolved["has_geo"]:
                geo_ok += 1
            else:
                geo_blank += 1

        self.issue_ids.unlink()
        self.env["res.partner.import.wizard.issue"].create(
            [
                {
                    "wizard_id": self.id,
                    "category": category,
                    "value": value,
                    "row_count": count,
                }
                for (category, value), count in issue_counter.items()
            ]
        )
        self.total_rows = len(raw_rows)
        self.rows_skipped = skipped
        self.rows_to_insert = len(raw_rows) - skipped
        self.geo_ok_count = geo_ok
        self.geo_blank_count = geo_blank
        self.state = "preview"
        return self._reopen()

    def action_back_to_upload(self):
        self.ensure_one()
        self.issue_ids.unlink()
        self.state = "upload"
        return self._reopen()

    # ------------------------------------------------------------------
    # Paso 2 -> 3: insertar por SQL (bypass ORM) lo que no esté excluido.
    # ------------------------------------------------------------------

    def action_confirm(self):
        self.ensure_one()
        raw_rows = self._parse_workbook(base64.b64decode(self.file))
        lookups = self._build_lookups()
        bejerman_counts = self._count_bejerman_codes(raw_rows)

        rows_by_column = []
        skipped = 0
        for raw in raw_rows:
            resolved = self._resolve_row(raw, lookups, bejerman_counts)
            if resolved["skip"]:
                skipped += 1
                continue
            rows_by_column.append(resolved["values"])

        new_ids = self._bulk_insert_partners(rows_by_column)

        self.inserted_count = len(new_ids)
        self.rows_skipped = skipped
        self.state = "done"
        return self._reopen()

    def _bulk_insert_partners(self, rows_by_column):
        if not rows_by_column:
            return []

        values_list = [
            tuple(row[column] for column in RES_PARTNER_COLUMNS)
            for row in rows_by_column
        ]
        columns_sql = ", ".join(RES_PARTNER_COLUMNS)
        sql = f"INSERT INTO res_partner ({columns_sql}) VALUES %s RETURNING id"
        new_ids = [
            row[0]
            for row in self.env.cr.execute_values(
                sql, values_list, page_size=5000, fetch=True
            )
        ]

        # commercial_partner_id es un related/compute store=True (apunta a
        # sí mismo cuando no hay parent_id, que es siempre el caso de este
        # import): el INSERT crudo no lo completa, y si queda NULL rompe
        # cualquier filtro/reporte que dependa de él.
        self.env.cr.execute(
            "UPDATE res_partner SET commercial_partner_id = id WHERE id = ANY(%s)",
            (new_ids,),
        )
        return new_ids

    # ------------------------------------------------------------------
    # Lectura del Excel
    # ------------------------------------------------------------------

    def _parse_workbook(self, content):
        content = self._strip_autofilters(content)
        try:
            workbook = openpyxl.load_workbook(
                io.BytesIO(content), read_only=True, data_only=True
            )
        except Exception as exc:  # pylint: disable=broad-except # noqa: BLE001
            raise UserError(
                _("No se pudo leer el archivo como Excel: %s", exc)
            ) from exc

        # row_limit corta la lectura apenas se junta esa cantidad de filas
        # (no sigue leyendo el resto del archivo): sirve para probar rápido
        # sin esperar a parsear las ~158k filas reales cada vez.
        limit = self.row_limit or None
        rows = []
        for sheet in workbook.worksheets:
            if limit and len(rows) >= limit:
                break
            header_row = next(
                sheet.iter_rows(min_row=1, max_row=1, values_only=True), None
            )
            if not header_row:
                continue

            col_index = {}
            for i, header in enumerate(header_row):
                key = HEADER_MAP.get(_normalize(header))
                if key and key != "_ignored":
                    col_index[key] = i

            missing = [k for k in REQUIRED_KEYS if k not in col_index]
            if missing:
                raise UserError(
                    _(
                        "Hoja '%(sheet)s': faltan columnas: %(missing)s",
                        sheet=sheet.title,
                        missing=", ".join(missing),
                    )
                )

            for row_number, raw_row in enumerate(
                sheet.iter_rows(min_row=2, values_only=True), start=2
            ):
                if limit and len(rows) >= limit:
                    break
                if all(cell is None for cell in raw_row):
                    continue
                row = {key: raw_row[idx] for key, idx in col_index.items()}
                row["_sheet"] = sheet.title
                row["_row_number"] = row_number
                rows.append(row)
        return rows

    @staticmethod
    def _strip_autofilters(content):
        try:
            with zipfile.ZipFile(io.BytesIO(content)) as zin:
                names = zin.namelist()
                buffer = io.BytesIO()
                with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zout:
                    for name in names:
                        data = zin.read(name)
                        if name.startswith("xl/worksheets/sheet") and name.endswith(
                            ".xml"
                        ):
                            data = _AUTOFILTER_RE.sub(b"", data)
                        zout.writestr(name, data)
                return buffer.getvalue()
        except zipfile.BadZipFile:
            # No es un .xlsx real (zip inválido): se deja pasar tal cual
            # para que openpyxl tire su propio error, más claro que este.
            return content

    # Prefijo para el código Bejerman sintético (ver _effective_bejerman_value):
    # los códigos Bejerman reales son siempre numéricos, así que un valor con
    # este prefijo no puede coincidir jamás con uno real de otro cliente.
    NO_BEJERMAN_PREFIX = "SC-"

    @classmethod
    def _effective_bejerman_value(cls, raw):
        """Código Bejerman a usar para la fila: la columna B (codigo
        bejerman) tal cual si está seteada y no es 0; si no, se cae al
        código de cliente de la columna A (customer_code), prefijado con
        NO_BEJERMAN_PREFIX.

        El prefijo es necesario porque "código de cliente" y "código
        Bejerman" son dos numeraciones independientes que se solapan en
        rango: sin prefijo, dos clientes sin relación entre sí pueden
        terminar con el mismo valor efectivo (el código de cliente de uno
        coincide con el código Bejerman real de otro) y _count_bejerman_codes
        los marca como duplicados, excluyendo ambas filas."""
        value = str(raw.get("codigo_bejerman") or "").strip()
        if value and value != "0":
            return value
        customer_code = str(raw.get("customer_code") or "").strip()
        return f"{cls.NO_BEJERMAN_PREFIX}{customer_code}" if customer_code else ""

    @classmethod
    def _count_bejerman_codes(cls, raw_rows):
        """Cuenta ocurrencias del código Bejerman efectivo (ver
        _effective_bejerman_value) en TODO el archivo (ambas hojas): el que
        aparezca más de una vez se excluye completo en _resolve_row, ver
        comentario ahí."""
        return Counter(
            value
            for value in (cls._effective_bejerman_value(raw) for raw in raw_rows)
            if value
        )

    # ------------------------------------------------------------------
    # Resolución de FKs (una consulta por modelo, cacheada en memoria)
    # ------------------------------------------------------------------

    def _build_lookups(self):
        country_map = {
            _normalize(c["name"]): c["id"]
            for c in self.env["res.country"].search_read([], ["name"])
        }
        argentina = self.env["res.country"].search([("code", "=", "AR")], limit=1)
        state_map = (
            {
                _normalize(s["name"]): s["id"]
                for s in self.env["res.country.state"].search_read(
                    [("country_id", "=", argentina.id)], ["name"]
                )
            }
            if argentina
            else {}
        )
        # No usa el módulo core "l10n_latam_base" (desinstalado en esta
        # instancia): el fork propio l10n_ar_eynes tiene su propio modelo
        # para el tipo de documento (DNI/CUIT/CUIL).
        doc_type_map = {
            _normalize(d["name"]): d["id"]
            for d in self.env["res.document.type"].search_read([], ["name"])
        }
        # Tampoco existe "l10n_ar.afip.responsibility.type": "Posición
        # fiscal" en l10n_ar_eynes es el campo estándar de Odoo/account
        # (property_account_position_id -> account.fiscal.position),
        # sembrado con los valores AFIP (Consumidor Final, IVA RI, etc.)
        # en l10n_ar_eynes/data/fiscal_position_data.xml.
        afip_map = {
            _normalize(a["name"]): a["id"]
            for a in self.env["account.fiscal.position"].search_read([], ["name"])
        }
        client_type_map = {
            _normalize(c["description"]): c["id"]
            for c in self.env["client.type"].search_read([], ["description"])
        }
        # Para no duplicar contactos si el wizard se corre más de una vez
        # (mismo Excel, o uno que se superpone): se saltea una fila si su
        # codigo_bejerman ya existe en la base; si no tiene, se cae a
        # buscar por vat (CUIT/CUIL/DNI ya sin puntos ni guiones). Una
        # sola query por cada uno, no una por fila.
        self.env.cr.execute(
            "SELECT codigo_bejerman FROM res_partner WHERE codigo_bejerman IS NOT NULL"
        )
        existing_bejerman = {row[0] for row in self.env.cr.fetchall()}
        self.env.cr.execute("SELECT vat FROM res_partner WHERE vat IS NOT NULL")
        existing_vat = {row[0] for row in self.env.cr.fetchall()}
        return {
            "country": country_map,
            "state": state_map,
            "doc_type": doc_type_map,
            "afip": afip_map,
            "client_type": client_type_map,
            "company_id": self.env.company.id,
            "uid": self.env.uid,
            # Para validar el vat con el mismo criterio que
            # ResPartner.check_vat_ar (l10n_ar_eynes, @api.constrains
            # 'vat'/'country_id'/'document_type_id'): ese chequeo corre en
            # cualquier write() posterior sobre el contacto (aunque sea de
            # otro campo) y rechaza toda la operación si el vat quedó mal
            # formado, así que hay que dejarlo bien desde el INSERT.
            "cuit_id": self.env.ref(
                "l10n_ar_eynes.document_cuit", raise_if_not_found=False
            ).id,
            "cuil_id": self.env.ref(
                "l10n_ar_eynes.document_cuil", raise_if_not_found=False
            ).id,
            "dni_id": self.env.ref(
                "l10n_ar_eynes.document_dni", raise_if_not_found=False
            ).id,
            # Default cuando el Excel no trae tipo de documento ni CUIT (o
            # el CUIT viene pero es inválido): "Doc. (Otro)" + vat marcador
            # NOIMPORTADO, en vez de dejar el contacto sin documento.
            "doc_otro_id": self.env.ref(
                "l10n_ar_eynes.document_doc_otro", raise_if_not_found=False
            ).id,
            "existing_bejerman": existing_bejerman,
            "existing_vat": existing_vat,
        }

    # ------------------------------------------------------------------
    # Resolución de una fila
    # ------------------------------------------------------------------

    def _resolve_row(self, raw, lookups, bejerman_counts):
        issues = []

        nombre = str(raw.get("nombre") or "").strip()
        if not nombre:
            return {"skip": True, "issues": [("blank_name", "(nombre vacío)")]}

        codigo_bejerman, bejerman_issue = self._resolve_codigo_bejerman(
            raw, bejerman_counts
        )
        if bejerman_issue and bejerman_issue[0] == "duplicate_bejerman":
            return {"skip": True, "issues": [bejerman_issue]}
        if bejerman_issue:
            issues.append(bejerman_issue)

        country_id = self._lookup(
            raw.get("pais"), lookups["country"], issues, "unmatched_country"
        )
        state_id = self._lookup_state(raw.get("estado"), lookups["state"], issues)
        doc_type_id = self._lookup(
            raw.get("tipo_documento"), lookups["doc_type"], issues, "unmatched_doc_type"
        )
        afip_id = self._lookup(
            raw.get("posicion_fiscal"), lookups["afip"], issues, "unmatched_afip_resp"
        )
        client_type_id = self._lookup(
            raw.get("tipo_cliente"),
            lookups["client_type"],
            issues,
            "unmatched_client_type",
        )

        lat = self._normalize_coord(raw.get("lat"))
        lon = self._normalize_coord(raw.get("lon"))

        calle = str(raw.get("calle") or "").strip() or None
        numero = str(raw.get("numero") or "").strip() or None
        street = " ".join(part for part in (calle, numero) if part) or None

        # fecha_alta (ivess_partner_custom) guarda la fecha de alta real del
        # sistema origen; create_date queda para lo que realmente es: cuándo
        # se creó el registro en Odoo (el momento de este import).
        fecha_alta = self._parse_datetime(raw.get("fecha_alta"))
        now = fields.Datetime.now()

        # Misma normalización que ResPartner._parse_vat_string en
        # l10n_ar_eynes: guarda el vat sin puntos ni guiones. Hay que
        # limpiarlo ACÁ antes de validar el dígito verificador, si no un
        # CUIT/CUIL bien formado pero escrito "30-71767557-2" se descarta
        # por longitud/formato en vez de validarse correctamente.
        tipo_documento_raw = str(raw.get("tipo_documento") or "").strip()
        vat_raw = str(raw.get("vat") or "").strip()

        vat = vat_raw.replace(".", "").replace("-", "") or None
        invalid_vat = False
        if vat and not self._check_vat_ar(
            vat,
            doc_type_id,
            lookups["cuit_id"],
            lookups["cuil_id"],
            lookups["dni_id"],
        ):
            issues.append(("invalid_vat_format", vat))
            invalid_vat = True
            vat = None

        # Sin tipo de documento ni CUIT (columnas P/Q vacías o en 0), o con
        # un CUIT/DNI que no pasa la validación del dígito verificador: se
        # completa con "Doc. (Otro)" + NOIMPORTADO en vez de dejar el
        # contacto sin documento.
        doc_and_vat_blank = tipo_documento_raw in ("", "0") and vat_raw in ("", "0")
        if doc_and_vat_blank or invalid_vat:
            doc_type_id = lookups["doc_otro_id"]
            vat = "NOIMPORTADO"

        # Evita duplicar contactos si el wizard se corre más de una vez:
        # se saltea la fila si su identificador ya existe en la base
        # (codigo_bejerman primero; si la fila no tiene, se cae a vat). El
        # marcador NOIMPORTADO no es un identificador real (lo comparten
        # todos los contactos sin CUIT), así que se lo excluye de este
        # chequeo para no saltear de más en corridas futuras.
        if codigo_bejerman and codigo_bejerman in lookups["existing_bejerman"]:
            return {"skip": True, "issues": [("already_imported", codigo_bejerman)]}
        if (
            not codigo_bejerman
            and vat
            and vat != "NOIMPORTADO"
            and vat in lookups["existing_vat"]
        ):
            return {"skip": True, "issues": [("already_imported", vat)]}

        values = {
            "name": nombre,
            "is_company": _normalize(raw.get("tipo_empresa")) == "empresa",
            "email": str(raw.get("email") or "").strip() or None,
            "phone": str(raw.get("phone") or "").strip() or None,
            "mobile_number": str(raw.get("mobile") or "").strip() or None,
            "street": street,
            # street_name/street_number (base_address_extended) son compute
            # store=True A PARTIR de street en el módulo, pero ese compute
            # corre write() registro por registro (157k veces atravesando
            # toda la cadena de módulos que heredan res.partner.write() es
            # carísimo, tarda más de una hora). Como el Excel ya trae Calle
            # y Numero separados, se setean directo sin pasar por el compute.
            "street_name": calle,
            "street_number": numero,
            "city": str(raw.get("ciudad") or "").strip() or None,
            "zip": str(raw.get("zip") or "").strip() or None,
            "country_id": country_id,
            "state_id": state_id,
            "vat": vat,
            "document_type_id": doc_type_id,
            # property_account_position_id es "company_dependent": Odoo lo
            # guarda como jsonb {"<company_id>": <fiscal_position_id>}, no
            # como integer (ver comentario en RES_PARTNER_COLUMNS). No se
            # puede setear vía write() del ORM en lote: el write() propio
            # de l10n_ar_eynes para res.partner asume un solo registro a
            # la vez (self.name sin ensure_one) y explota con "Expected
            # singleton" al escribir sobre más de un partner junto.
            "property_account_position_id": (
                json.dumps({str(lookups["company_id"]): afip_id}) if afip_id else None
            ),
            "partner_type_id": client_type_id,
            "customer_code": str(raw.get("customer_code") or "").strip() or None,
            "codigo_bejerman": codigo_bejerman,
            "fecha_alta": fecha_alta,
            "is_important_client": self._to_bool(raw.get("cliente_importante")),
            "requiere_comprobante": self._to_bool(raw.get("requiere_comprobante")),
            "has_special_price": self._to_bool(raw.get("precios_especiales")),
            "comment": self._build_comment(
                raw.get("etiqueta"), raw.get("obs_direccion"), raw.get("notas")
            ),
            "partner_latitude": lat,
            "partner_longitude": lon,
            "create_date": now,
            "write_date": now,
            "create_uid": lookups["uid"],
            "write_uid": lookups["uid"],
            "company_id": lookups["company_id"],
            "active": True,
            "type": "contact",
            "autopost_bills": "ask",
            "group_rfq": "default",
            "group_on": "default",
        }
        return {
            "skip": False,
            "issues": issues,
            "values": values,
            "has_geo": lat is not None and lon is not None,
        }

    @staticmethod
    def _check_vat_ar(vat, doc_type_id, cuit_id, cuil_id, dni_id):
        """Réplica de ResPartner.check_vat_ar (l10n_ar_eynes): ese método
        corre como @api.constrains('vat', 'country_id', 'document_type_id')
        y rechaza CUALQUIER write() posterior sobre el contacto (aunque sea
        de un campo no relacionado) si el vat quedó mal formado. Hay que
        validar antes de insertar, no después."""
        if doc_type_id in (cuit_id, cuil_id):
            if len(vat) != 11:
                return False
        elif doc_type_id == dni_id:
            if len(vat) > 8:
                return False
        else:
            return True
        if not vat.isdigit():
            return False
        if doc_type_id in (cuit_id, cuil_id):
            check_list = [5, 4, 3, 2, 7, 6, 5, 4, 3, 2]
            total = sum(int(vat[i]) * check_list[i] for i in range(10))
            remainder = 11 - total % 11
            if remainder == 11:
                remainder = 0
            if remainder == 10:
                remainder = 9
            return remainder == int(vat[10])
        return True

    @classmethod
    def _resolve_codigo_bejerman(cls, raw, bejerman_counts):
        value = cls._effective_bejerman_value(raw)
        if not value:
            return None, None
        if bejerman_counts[value] > 1:
            return None, ("duplicate_bejerman", value)
        return value, None

    @staticmethod
    def _lookup(raw_value, mapping, issues, category):
        text = str(raw_value or "").strip()
        if not text:
            return None
        result = mapping.get(_normalize(text))
        if not result:
            issues.append((category, text))
        return result

    @staticmethod
    def _lookup_state(raw_value, mapping, issues):
        text = str(raw_value or "").strip()
        if not text:
            return None
        key = STATE_COUNTRY_SUFFIX_RE.sub("", _normalize(text)).strip()
        result = mapping.get(key)
        if not result:
            issues.append(("unmatched_state", text))
        return result

    @staticmethod
    def _to_bool(raw_value):
        return str(raw_value or "").strip().upper() == "TRUE"

    @staticmethod
    def _normalize_coord(raw_value):
        """El Excel origen trae latitud/longitud sin punto decimal (ej.
        "-34641179" en vez de "-34.641179"): Argentina cae siempre en 1-2
        dígitos de parte entera para lat/lon, así que se reconstruye tomando
        los primeros 2 dígitos como parte entera y el resto como decimales.
        Verificado contra las ~155k filas con dato: los prefijos son
        consistentemente 34/35 (latitud) y 57/58 (longitud)."""
        text = str(raw_value or "").strip()
        if text in ("", "0"):
            return None
        if "." in text:
            try:
                return float(text)
            except ValueError:
                return None
        negative = text.startswith("-")
        digits = text.lstrip("-")
        if len(digits) < 2:
            return None
        value = float(f"{digits[:2]}.{digits[2:] or '0'}")
        return -value if negative else value

    @staticmethod
    def _parse_datetime(raw_value):
        if raw_value is None or raw_value == "":
            return None
        if isinstance(raw_value, datetime):
            return raw_value
        text = str(raw_value).strip()
        for fmt in ("%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
        return None

    @staticmethod
    def _build_comment(etiqueta, obs_direccion, notas):
        parts = []
        etiqueta = str(etiqueta or "").strip()
        if etiqueta:
            parts.append(_("Etiqueta origen: %s", etiqueta))
        obs_direccion = str(obs_direccion or "").strip()
        if obs_direccion:
            parts.append(_("Observaciones de dirección: %s", obs_direccion))
        notas = str(notas or "").strip()
        if notas:
            parts.append(notas)
        if not parts:
            return None
        joined = "<br/>".join(str(escape(part)) for part in parts)
        return f"<p>{joined}</p>"

    def _reopen(self):
        return {
            "type": "ir.actions.act_window",
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
        }
